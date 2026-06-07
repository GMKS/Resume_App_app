from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_DIR = Path(r"c:\Resume_App_app\release-artifacts\qa")
OUTPUT_FILE = OUTPUT_DIR / "resumix_ai_enterprise_qa_test_suite_700.xlsx"
BUILD_VERSION = "RESUMIX-AI Android RC 2026.05.28"

MANDATORY_COLUMNS = [
    "Test Case ID",
    "Test Type",
    "Module Name",
    "Feature Name",
    "Test Scenario",
    "Preconditions",
    "Test Steps",
    "Test Data",
    "Expected Result",
    "Actual Result",
    "Status",
    "Priority",
    "Severity",
    "Environment",
    "Device/Platform",
    "Build Version",
    "Remarks/Comments",
]

ALL_REQUIRED_TEST_TYPES = [
    "Functional Testing",
    "Exploratory Testing",
    "Regression Testing",
    "Smoke Testing",
    "Sanity Testing",
    "UI Testing",
    "UX Testing",
    "Responsive Testing",
    "Compatibility Testing",
    "Cross-device Testing",
    "End-to-End Testing",
    "Integration Testing",
    "System Testing",
    "Acceptance Testing",
    "Validation Testing",
    "Security Testing",
    "Performance Testing",
    "Stability Testing",
    "Accessibility Testing",
    "Localization Testing",
    "Interrupt Testing",
    "Installation/Uninstallation Testing",
    "Recovery Testing",
    "API Testing",
    "Navigation Testing",
    "Data Persistence Testing",
    "AI Feature Testing",
    "PDF Export Testing",
    "Cloud Backup/Restore Testing",
]

DEVICE_POOL = [
    "Google Pixel 8 / Android 14",
    "Samsung Galaxy A52 / Android 13",
    "OnePlus 11 / Android 14",
    "Xiaomi Redmi Note 12 / Android 13",
    "Motorola G54 / Android 12",
    "Samsung Tab S9 / Android 14",
    "Samsung Z Fold 5 / Android 14",
    "Android Go reference device / Android 12",
]

MODULE_PALETTE = [
    "DDEBF7",
    "E2F0D9",
    "FCE4D6",
    "FFF2CC",
    "E4DFEC",
    "D9EAF7",
    "F4CCCC",
    "D0E0E3",
    "FCE5CD",
    "D9D2E9",
    "D9EAD3",
    "CFE2F3",
]

TEST_TYPE_PALETTE = [
    "EAF2F8",
    "E8F6F3",
    "FEF9E7",
    "FDEDEC",
    "EBF5FB",
    "F4ECF7",
    "E8F8F5",
    "FCF3CF",
    "FDEBD0",
    "F5EEF8",
]

PRIORITY_COLORS = {
    "Critical": "F4CCCC",
    "High": "FCE5CD",
    "Medium": "FFF2CC",
    "Low": "D9EAD3",
}

SEVERITY_COLORS = {
    "Blocker": "E06666",
    "Critical": "F4CCCC",
    "Major": "F9CB9C",
    "Moderate": "FFE599",
    "Minor": "D9EAD3",
}

STATUS_COLORS = {
    "Not Executed": "D9EAF7",
}

BASE_PRIORITY = {
    "lifecycle": "Critical",
    "navigation": "Medium",
    "home": "Medium",
    "resume": "High",
    "auth": "Critical",
    "editor": "High",
    "templates": "Medium",
    "preview": "Critical",
    "export": "Critical",
    "ai": "High",
    "career": "Medium",
    "profile": "Medium",
    "sync": "Critical",
    "settings": "High",
    "subscription": "Critical",
    "compliance": "Critical",
}


@dataclass(frozen=True)
class Feature:
    module: str
    feature: str
    screen: str
    objective: str
    sample_data: str
    invalid_data: str
    boundary_data: str
    integration: str
    category: str


FEATURES = [
    Feature(
        "App Lifecycle",
        "Google Play installation and first launch",
        "Play Internal Test install flow and splash screen",
        "install the Android build from a Play distribution and reach the first usable screen",
        "Fresh install from Google Play Internal Test track on Android 14",
        "Install package with low storage or interrupted download",
        "Install on a nearly full device with 500 MB free space",
        "Google Play package delivery, Firebase initialization, and app bootstrap services",
        "lifecycle",
    ),
    Feature(
        "App Lifecycle",
        "Update existing install and local data migration",
        "app update path from a previously installed build",
        "update the app without losing resumes, settings, subscriptions, or cached preview assets",
        "Upgrade from a prior production-like build containing 5 resumes and active settings",
        "Update while the app is backgrounded and storage is constrained",
        "Upgrade across two skipped versions with old Hive data present",
        "Android package manager, Hive schema loading, SharedPreferences migration, and Firebase bootstrap",
        "lifecycle",
    ),
    Feature(
        "App Lifecycle",
        "Uninstall, reinstall, and data cleanup behavior",
        "device app management and first-run relaunch",
        "remove the app cleanly and confirm local data is recreated only when intended",
        "Uninstall then reinstall on the same device using the same Google account",
        "Reinstall after an interrupted uninstall or corrupted cache scenario",
        "Repeated install/uninstall cycles across low storage states",
        "Android package manager, local storage lifecycle, and optional cloud restore entry points",
        "lifecycle",
    ),
    Feature(
        "App Lifecycle",
        "Cold start, splash, and startup recovery",
        "splash and app initialization sequence",
        "recover cleanly from cold starts, force stops, and partial startup failures",
        "Launch after force stop with network enabled and valid cached user data",
        "Launch with corrupted cache, revoked network, or failing Firebase init",
        "Cold start after 30 days idle with large cached preview and resume data",
        "startup guards, Firebase init, config loading, and app shell navigation",
        "lifecycle",
    ),
    Feature(
        "Onboarding",
        "First-run onboarding carousel",
        "onboarding screen",
        "educate a new user and transition correctly into the app shell",
        "First launch with no prior onboarding completion flag",
        "Swipe rapidly, rotate device, or skip mid-animation",
        "Very small screen and large accessibility font while onboarding is shown",
        "local onboarding state, navigation shell, and app theme loading",
        "navigation",
    ),
    Feature(
        "Dashboard",
        "Bottom navigation and dashboard shell",
        "main dashboard with Home, Resumes, Career Tools, Portfolio, and Profile tabs",
        "move between major sections without losing the current state or triggering incorrect back behavior",
        "Tab switching across all five tabs with active content in each tab",
        "Repeated fast tapping, back presses, and restored state after process restart",
        "Use navigation on a foldable and tablet in portrait and landscape",
        "navigation state providers, nested tab screens, and route restoration",
        "navigation",
    ),
    Feature(
        "Home",
        "Home quick actions, summary cards, and stats widgets",
        "home screen",
        "surface actionable shortcuts and correct resume/job metrics",
        "Home screen with multiple resumes, active subscription, and job tracker entries",
        "No resumes, empty stats, or stale cached counts",
        "Long user names and large data counts on compact screens",
        "resume provider, job tracker storage, and shared summary widgets",
        "home",
    ),
    Feature(
        "Resume Management",
        "My Resumes list and header actions",
        "My Resumes screen",
        "present resume inventory, sorting, and create-entry actions clearly",
        "Existing account with 10 resumes, mixed completion percentages, and thumbnails",
        "No resumes, partially saved drafts, or broken preview cache",
        "Long resume titles, mixed templates, and tablet layout",
        "resume list provider, storage service, and preview/thumbnail dependencies",
        "resume",
    ),
    Feature(
        "Resume Management",
        "Create, duplicate, and delete resume workflows",
        "resume list and contextual actions",
        "create new resumes and safely duplicate or delete existing ones",
        "Create a resume from empty state, duplicate a completed resume, delete a draft resume",
        "Delete during low storage or duplicate when free-plan limits are reached",
        "Duplicate a resume with many sections, images, and custom sections",
        "local storage service, subscription checks, and navigation refresh",
        "resume",
    ),
    Feature(
        "Resume Management",
        "Resume cards, empty states, and completion indicators",
        "resume cards and empty-state widgets",
        "show meaningful empty states and accurate completion percentages for each resume",
        "Resume cards with 0%, 64%, and 100% completion states",
        "Corrupted completion inputs or missing preview metadata",
        "Large content cards on narrow phones and tablets",
        "completion calculators, resume card widgets, and empty-state presentation",
        "resume",
    ),
    Feature(
        "Authentication",
        "Phone OTP request",
        "phone sign-in form",
        "request an OTP only for valid phone numbers and surface safe error messages",
        "Valid phone numbers with country code and active network",
        "Malformed phone numbers, banned prefixes, blank input, or offline device",
        "Longest supported international number with spaces and separators",
        "OTP send endpoint, auth throttling, and disclosure messaging",
        "auth",
    ),
    Feature(
        "Authentication",
        "OTP verification and session establishment",
        "OTP verification screen",
        "verify a correct code, establish a session, and handle expiry or mismatch safely",
        "Valid OTP code returned from test backend for a signed-in phone",
        "Expired OTP, incorrect code, too many attempts, or stale session",
        "OTP entry with paste, auto-fill, and delayed network response",
        "OTP verify endpoint, Firebase/custom auth session handling, and masked-phone state",
        "auth",
    ),
    Feature(
        "Authentication",
        "Google social sign-in",
        "social sign-in buttons and auth handoff",
        "sign in with Google and return to the app with a valid session and profile data",
        "Firebase Google auth enabled with tester Gmail on a Play-installed build",
        "User cancels account chooser, missing SHA config, or network/auth provider failure",
        "Sign in on a device with multiple Google accounts and work profile enabled",
        "Google Play Services, Firebase Auth, and profile/session persistence",
        "auth",
    ),
    Feature(
        "Authentication",
        "Session persistence, re-entry, and logout",
        "app launch with existing user session",
        "restore the right signed-in state and log out cleanly without residue",
        "Launch with valid prior OTP or Google-auth session stored locally",
        "Expired tokens, revoked provider access, or partial local session state",
        "Session recovery after device restart and app process kill",
        "auth state providers, local session store, and profile refresh",
        "auth",
    ),
    Feature(
        "Resume Editor",
        "Editor toolbar, autosave, and section shortcuts",
        "resume editor screen toolbar",
        "save reliably while navigating to preview, ATS, reorder, theme, and formatting tools",
        "Open a resume with multiple completed sections and active autosave",
        "Switch sections rapidly during a background save or after connectivity loss",
        "Toolbar usage on narrow phones with horizontal scrolling enabled",
        "resume provider, autosave logic, preview route, ATS route, and template route",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "Personal information editor",
        "personal info screen",
        "capture and validate personal details, contact info, links, and optional photo metadata",
        "Full name, email, phone, address, LinkedIn, GitHub, website, and job title",
        "Invalid email, broken links, blank required fields, and unsupported image input",
        "Maximum realistic address length and international phone formatting",
        "resume storage, preview rendering, and export field mapping",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "Professional summary editor",
        "summary screen",
        "save concise summaries and keep formatting stable across preview and export",
        "150-word summary tailored to a QA Engineer persona",
        "Empty summary, multi-line pasted text with stray whitespace, or unsupported characters",
        "Long summary near expected upper boundary with punctuation and symbols",
        "resume storage, AI rewrite dependencies, and template preview text layout",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "Experience editor",
        "experience screen",
        "capture work history, date ranges, achievements, and current-role logic accurately",
        "Two jobs with bullets, achievements, current-role toggle, and location fields",
        "End date before start date, blank company, or overlapping invalid ranges",
        "10+ achievements and long bullet sets for one experience record",
        "resume storage, preview grouping, AI enhancement, and export layout",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "Education editor",
        "education screen",
        "store degrees, institutions, grades, and date ranges consistently",
        "Two education entries with grades, location, and active-study flag",
        "Invalid date range, blank institution, or unsupported grade format",
        "Very long institution names and multi-degree entries",
        "resume storage and template-specific education blocks",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "Skills editor",
        "skills screen",
        "add, edit, categorize, and rate skills without duplicate or broken states",
        "Add technical and soft skills with multiple proficiency levels",
        "Blank skill, repeated skill name, or invalid category mapping",
        "50+ skills split across categories for a senior profile",
        "resume storage, ATS analyzer, and template-side skill rendering",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "Projects editor",
        "projects screen",
        "maintain project titles, links, technologies, and dates for preview and export",
        "Three projects with URLs, technologies, and date spans",
        "Malformed URL, empty title, or missing description",
        "Long technology stacks and projects with no end date",
        "resume storage, portfolio sync, preview widgets, and PDF export blocks",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "Certifications editor",
        "certifications screen",
        "capture certification metadata accurately for premium and ATS templates",
        "Certification name, issuer, issue date, expiry date, credential ID, and URL",
        "Expiry before issue date, blank certificate name, or bad credential URL",
        "Long certification names and multiple credentials from the same issuer",
        "resume storage, template mapping, and ATS keyword extraction",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "Languages editor",
        "languages screen",
        "manage languages and proficiency levels cleanly in all templates",
        "Add English, Hindi, and Spanish with different proficiencies",
        "Blank language name or duplicate entries with conflicting proficiency",
        "10+ languages and long proficiency labels for localization stress",
        "resume storage, preview tags, and export rendering",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "Built-in custom section editor",
        "custom section screen",
        "edit template-provided optional sections without losing the configured section identity",
        "Startup or professional-role template with optional sections enabled",
        "Missing config title, empty items, or unsupported section ID mapping",
        "Long optional section content with multiple items and bullet text",
        "template section config, storage, and preview/PDF title mapping",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "User custom section create, edit, rename, and delete",
        "user custom section screen and create sheet",
        "let users maintain custom sections with unique titles and durable content",
        "Create Awards, Publications, and Open Source Contributions sections",
        "Duplicate title, blank title, empty content confirmation, or delete cancellation",
        "Multiple custom sections with long names and multiline entries",
        "custom section model, serialization, preview/PDF binding, and cloud restore",
        "editor",
    ),
    Feature(
        "Resume Editor",
        "Section reorder flow",
        "Reorder Sections bottom sheet",
        "persist section order correctly across built-in and user-defined sections",
        "Resume with standard sections plus multiple custom sections having unique titles",
        "Legacy order keys, duplicated labels, or reorder interrupted before apply",
        "Reorder across 15+ sections on a compact device",
        "section-order persistence, custom-section mapping, preview order, and export order",
        "editor",
    ),
    Feature(
        "Templates",
        "Template selection",
        "template selection screen",
        "apply templates safely and keep existing resume data intact",
        "Switch between modern, ATS, startup, executive, designer, and premium templates",
        "Select locked templates on free plan or switch during a preview refresh",
        "Rapidly change templates across 10 options with loaded custom sections",
        "template normalization, premium gating, and preview refresh pipeline",
        "templates",
    ),
    Feature(
        "Templates",
        "Theme, font, and layout customization",
        "theme, format, and typography controls",
        "apply color, font, and layout selections without corrupting previews or exports",
        "Switch colors, font families, and layout style on an existing resume",
        "Choose locked font/layout on free plan or unsupported combo",
        "Very long resume content rendered with multiple fonts and compact layouts",
        "theme service, font mapping, preview cache, and PDF service",
        "templates",
    ),
    Feature(
        "Templates",
        "Template preview fidelity across variants",
        "template-specific preview widgets",
        "render each template accurately with the resume data provided to it",
        "Resume with photo, custom sections, long bullets, links, and mixed section density",
        "Template with missing optional sections, empty lists, or stale cached render",
        "Long content on two-column, timeline, and card-heavy templates",
        "template support mappers, preview widgets, and color/font settings",
        "templates",
    ),
    Feature(
        "Preview",
        "Live preview generation and refresh",
        "preview screen",
        "generate preview content promptly and reflect the latest saved resume state",
        "Edited resume with recent changes in multiple sections and template settings",
        "Open preview during background save, missing image, or stale provider state",
        "Preview a data-heavy resume with 20+ sections/items and image assets",
        "preview services, cache invalidation, provider refresh, and watermark overlay",
        "preview",
    ),
    Feature(
        "Preview",
        "PDF export generation",
        "preview export action",
        "generate a valid PDF file that reflects the selected template and the latest data",
        "Export a complete resume with custom sections, links, and image/photo fields",
        "Export while storage permission or share target is unavailable",
        "Generate PDF for very long resumes spanning multiple pages and tables",
        "preview PDF service, template-specific PDF renderers, and file IO",
        "export",
    ),
    Feature(
        "Preview",
        "Share, open, and print exported documents",
        "export share/open flow",
        "hand off exported documents to Android intents and external viewers cleanly",
        "Share PDF to Gmail, Drive, and local file viewers from a completed preview",
        "No compatible share target, printer unavailable, or share interrupted mid-flow",
        "Large exported file on a low-memory device",
        "shareable file helpers, Android intents, print/open targets, and file validation",
        "export",
    ),
    Feature(
        "Preview",
        "Export file validation, watermark, and freemium rules",
        "preview export validation path",
        "enforce export readiness rules, watermark behavior, and file integrity expectations",
        "Free-plan and premium-plan exports from the same resume and template",
        "Attempt export with invalid state, unsupported file path, or gated premium configuration",
        "Repeated export generation with long filenames and duplicate timestamps",
        "free-plan checks, PDF watermark behavior, file validation helpers, and storage IO",
        "export",
    ),
    Feature(
        "AI Tools",
        "AI resume generator",
        "AI resume generator screen",
        "build a structured resume from user prompts and seed data",
        "Generate a QA Engineer resume from role, skills, and experience inputs",
        "Missing prompt fields, malformed AI response, or provider timeout",
        "Rich prompt with multilingual text, numbers, and special characters",
        "AI provider request/response handling and resume model mapping",
        "ai",
    ),
    Feature(
        "AI Tools",
        "AI content enhancer",
        "AI content enhancer screen",
        "enhance existing resume content while preserving user control over saved changes",
        "Enhance summary and experience bullets for an existing resume",
        "Blank source section, revoked API key, or provider rate limit",
        "Large source text blocks pasted from external sources",
        "AI provider calls, local resume updates, and diff/apply behavior",
        "ai",
    ),
    Feature(
        "AI Tools",
        "AI bullet generator",
        "AI bullet generator screen",
        "generate strong bullet points from job/activity inputs and reuse them in editor flows",
        "Generate bullets from role, action, impact, and technology inputs",
        "Missing mandatory prompt fields or empty AI output",
        "Long role descriptions and mixed numeric impact details",
        "AI provider integration and resume editor copy/apply workflow",
        "ai",
    ),
    Feature(
        "AI Tools",
        "AI resume rewrite",
        "AI resume rewrite screen",
        "rewrite selected sections while keeping the underlying resume structure valid",
        "Rewrite summary, experience, and skills for a target job role",
        "Provider rejects payload or returns structurally invalid rewrite JSON",
        "Rewrite content containing multilingual text and special formatting",
        "AI request mapping, rewrite response parsing, and resume update flow",
        "ai",
    ),
    Feature(
        "AI Tools",
        "AI job tailor",
        "AI job tailor screen",
        "tailor resume content to a pasted job description without dropping critical sections",
        "Resume plus full job description and selected tailoring scope",
        "Missing job description, huge pasted payload, or partial AI response",
        "Tailoring with multiple custom sections and domain-specific keywords",
        "AI provider, custom section preservation, and resume merge logic",
        "ai",
    ),
    Feature(
        "ATS",
        "ATS analysis and optimization",
        "ATS analyzer and optimization screens",
        "score a resume, identify gaps, and drive actionable ATS improvements",
        "Analyze a resume against a target role and review optimization advice",
        "Open ATS tool without enough resume data or with malformed job description",
        "Large skill matrix and job description loaded with uncommon keywords",
        "ATS analyzers, scoring engine, keyword extraction, and optimization UI",
        "ai",
    ),
    Feature(
        "AI Tools",
        "LinkedIn import",
        "LinkedIn import screen",
        "parse imported profile data into the resume model without field loss",
        "Import structured LinkedIn text/profile content into a new resume",
        "Malformed import data, unsupported sections, or partial parsing",
        "Import a very long profile containing projects, awards, and certifications",
        "import parser, resume mapper, and editor navigation handoff",
        "ai",
    ),
    Feature(
        "AI Tools",
        "AI assistant and RAOE2 flows",
        "AI assistant and RAOE2 screens",
        "support iterative AI guidance and structured improvements across multiple resume sections",
        "Use AI assistant prompts and RAOE2 rewriting on an existing resume",
        "Empty prompt, unsupported tone, or provider/network failure",
        "Multiple sequential AI actions without closing the app",
        "AI orchestration services, provider auth, and local resume state transitions",
        "ai",
    ),
    Feature(
        "AI Tools",
        "Resume roast and score checker",
        "resume roast screen",
        "generate critical feedback without blocking edits or corrupting resume data",
        "Request a score and roast for a completed resume",
        "Low-data resume, provider outage, or rate-limited account",
        "Roast a resume with very long descriptions and many custom sections",
        "AI review provider flow, score breakdown rendering, and non-destructive UX",
        "ai",
    ),
    Feature(
        "Career Tools",
        "Job tracker",
        "job tracker screen",
        "manage job applications, statuses, notes, reminders, and activities reliably",
        "Create applications with interviews, notes, and activity timelines",
        "Blank company/role, invalid dates, or duplicated applications",
        "200+ tracker entries with long notes and mixed statuses",
        "local storage, optional sync, reminders, and filtering/sorting widgets",
        "career",
    ),
    Feature(
        "Career Tools",
        "Career tools utilities",
        "cover letter, interview prep, job search, skill analyzer, career path, and articles screens",
        "provide stable supporting tools around the core resume workflow",
        "Access each tool from Career Tools tab and use default interactive actions",
        "Open a tool with incomplete source data or missing connectivity",
        "Long content cards, embedded links, and tool switching on tablets",
        "tool routing, shared widgets, AI helpers, and content/state providers",
        "career",
    ),
    Feature(
        "Portfolio",
        "Portfolio tab",
        "portfolio tab screen",
        "show portfolio/project content consistently and persist local edits",
        "Portfolio with projects, certificates, and links already saved locally",
        "Broken links, empty portfolio, or partially saved edits",
        "Long card content and image-heavy project entries",
        "portfolio local storage, resume project dependencies, and tab shell navigation",
        "profile",
    ),
    Feature(
        "Profile",
        "Profile, help, privacy, notifications, and legal screens",
        "profile tab and linked support/legal screens",
        "let users review profile data, support options, privacy text, and notification preferences cleanly",
        "Signed-in and signed-out profile states with support links available",
        "Broken support link, missing policy page, or disabled notifications state mismatch",
        "Large text, landscape layout, and localized policy headings",
        "profile storage, settings links, legal content screens, and notification state",
        "profile",
    ),
    Feature(
        "Cloud Sync",
        "Cloud backup, restore, and sync code flows",
        "settings backup and restore paths",
        "back up resumes safely, restore them correctly, and respect sync codes and newer data",
        "Local and cloud resumes with overlapping IDs and different updated timestamps",
        "Invalid sync code, offline restore, or stale cloud payload",
        "Restore across devices with large resumes and many custom sections",
        "Supabase/Firebase sync service, resume JSON mapping, and conflict resolution",
        "sync",
    ),
    Feature(
        "Settings",
        "Settings, privacy controls, and delete-all-data flow",
        "settings screen",
        "surface configuration, privacy disclosures, and irreversible data-deletion actions safely",
        "Settings with cloud sync configured, AI key configured, and resumes present locally",
        "Delete while cloud service fails, unsupported config values, or missing permissions",
        "Very long privacy copy and small-screen settings scrolling",
        "preferences store, secure key storage, data deletion services, and backup state",
        "settings",
    ),
    Feature(
        "Subscription",
        "Purchase, restore, and manage premium subscription",
        "subscription screen",
        "complete Android billing flows and enforce premium entitlement boundaries correctly",
        "Google Play tester account with active subscription products configured",
        "Unavailable billing service, canceled purchase, or restore failure",
        "Multiple product tiers and plan switching with expired prior entitlement",
        "Google Play Billing, entitlement state, restore flow, and manage-subscription links",
        "subscription",
    ),
    Feature(
        "Compliance",
        "Responsive, accessibility, localization, compatibility, and Play Store production readiness",
        "cross-app production readiness audit path",
        "verify the Android build is production-ready for diverse devices, accessibility needs, and Play policy review",
        "Production-like release candidate installed from Play Internal Test with all major modules enabled",
        "Policy disclosure mismatch, missing metadata, placeholder config, or unsupported locale/assistive setup",
        "Large fonts, TalkBack, RTL/long localized text, foldable posture changes, and low-memory devices",
        "Play billing, auth providers, export path, privacy disclosures, and release configuration",
        "compliance",
    ),
]


TEMPLATE_LABELS = [
    "Primary workflow",
    "Validation guardrail",
    "Negative/error path",
    "Boundary and real-world data",
    "Persistence and state recovery",
    "Integration and downstream impact",
    "Navigation and UI clarity",
    "Responsive, accessibility, compatibility, and localization",
    "Interruption, recovery, stability, and performance",
    "Regression and production readiness",
]

EXPLORATORY_LABELS = [
    "Explore the first-time and golden-path journey",
    "Explore invalid inputs, odd data, and misuse patterns",
    "Explore connectivity loss, interruptions, and recovery behavior",
    "Explore responsive, accessibility, and cross-device behavior",
]


def build_preconditions(feature: Feature, template_index: int) -> str:
    network_note = "Stable internet is available." if feature.category in {"auth", "ai", "sync", "subscription", "preview", "export", "compliance"} else "Device is available in a stable QA state."
    base = [
        f"Android test build {BUILD_VERSION} is installed.",
        network_note,
        f"Tester can reach {feature.screen}.",
    ]
    if template_index in {4, 8, 9}:
        base.append("Baseline data relevant to the feature is already saved and visible.")
    if feature.category in {"sync", "subscription", "auth", "ai"}:
        base.append(f"Required external dependency is configured: {feature.integration}.")
    return " ".join(base)


def numbered_steps(*steps: str) -> str:
    return "\n".join(f"{index}. {step}" for index, step in enumerate(steps, start=1))


def choose_device(feature_index: int, template_index: int, test_type: str) -> str:
    if test_type in {
        "Responsive Testing",
        "Compatibility Testing",
        "Cross-device Testing",
        "Accessibility Testing",
        "Localization Testing",
    }:
        return f"{DEVICE_POOL[(feature_index + template_index) % len(DEVICE_POOL)]} + {DEVICE_POOL[(feature_index + template_index + 3) % len(DEVICE_POOL)]}"
    return DEVICE_POOL[(feature_index + template_index) % len(DEVICE_POOL)]


def choose_environment(category: str, test_type: str) -> str:
    if test_type in {"API Testing", "Integration Testing", "Cloud Backup/Restore Testing"}:
        return "Android Staging environment with test backend services"
    if test_type in {"Performance Testing", "Stability Testing"}:
        return "Android QA performance lab"
    if test_type in {"Acceptance Testing", "Installation/Uninstallation Testing", "Security Testing"} or category in {"subscription", "compliance", "lifecycle"}:
        return "Pre-production release candidate / Google Play Internal Test"
    if category == "ai":
        return "AI feature QA environment with controlled provider/test credentials"
    return "Android QA regression lab"


def select_primary_type(category: str, feature_index: int) -> str:
    pools = {
        "lifecycle": ["Smoke Testing", "Installation/Uninstallation Testing", "Acceptance Testing"],
        "auth": ["End-to-End Testing", "Functional Testing", "Acceptance Testing"],
        "resume": ["Functional Testing", "Smoke Testing", "End-to-End Testing"],
        "editor": ["Functional Testing", "End-to-End Testing", "Smoke Testing"],
        "preview": ["Functional Testing", "End-to-End Testing", "Smoke Testing"],
        "export": ["PDF Export Testing", "End-to-End Testing", "Functional Testing"],
        "ai": ["AI Feature Testing", "Functional Testing", "Acceptance Testing"],
        "sync": ["Cloud Backup/Restore Testing", "End-to-End Testing", "Functional Testing"],
        "subscription": ["Acceptance Testing", "End-to-End Testing", "Functional Testing"],
        "compliance": ["Acceptance Testing", "Security Testing", "Compatibility Testing"],
    }
    default_pool = ["Functional Testing", "Smoke Testing", "Acceptance Testing"]
    pool = pools.get(category, default_pool)
    return pool[feature_index % len(pool)]


def select_integration_type(feature: Feature, feature_index: int) -> str:
    if feature.category == "auth":
        return ["API Testing", "Integration Testing"][feature_index % 2]
    if feature.category == "sync":
        return ["Cloud Backup/Restore Testing", "Integration Testing"][feature_index % 2]
    if feature.category == "ai":
        return ["AI Feature Testing", "API Testing", "Integration Testing"][feature_index % 3]
    if feature.category == "export":
        return ["PDF Export Testing", "Integration Testing"][feature_index % 2]
    if feature.category == "subscription":
        return ["API Testing", "Integration Testing", "Acceptance Testing"][feature_index % 3]
    return ["Integration Testing", "System Testing"][feature_index % 2]


def select_compatibility_type(feature_index: int) -> str:
    pool = [
        "Responsive Testing",
        "Compatibility Testing",
        "Cross-device Testing",
        "Accessibility Testing",
        "Localization Testing",
    ]
    return pool[feature_index % len(pool)]


def select_interrupt_type(feature_index: int) -> str:
    pool = [
        "Interrupt Testing",
        "Recovery Testing",
        "Stability Testing",
        "Performance Testing",
    ]
    return pool[feature_index % len(pool)]


def select_regression_type(feature: Feature, feature_index: int) -> str:
    if feature.category == "lifecycle":
        return ["Installation/Uninstallation Testing", "Regression Testing", "Sanity Testing"][feature_index % 3]
    if feature.category in {"auth", "sync", "subscription", "compliance"}:
        return ["Security Testing", "Regression Testing", "System Testing"][feature_index % 3]
    return ["Regression Testing", "Sanity Testing", "System Testing"][feature_index % 3]


def priority_for(feature: Feature, test_type: str, template_index: int) -> str:
    base = BASE_PRIORITY[feature.category]
    if test_type in {"Security Testing", "Cloud Backup/Restore Testing", "Installation/Uninstallation Testing", "PDF Export Testing"}:
        return "Critical" if base in {"Critical", "High"} else "High"
    if test_type in {"Accessibility Testing", "Localization Testing", "Responsive Testing", "Compatibility Testing"} and base == "Medium":
        return "Medium"
    if template_index == 8 and base == "Medium":
        return "High"
    return base


def severity_for(feature: Feature, test_type: str, template_index: int, priority: str) -> str:
    if test_type in {"Security Testing", "Installation/Uninstallation Testing"} and feature.category in {"auth", "subscription", "sync", "compliance", "lifecycle"}:
        return "Blocker"
    if priority == "Critical" and template_index in {0, 5, 9}:
        return "Critical"
    if test_type in {"Recovery Testing", "Data Persistence Testing", "PDF Export Testing", "Cloud Backup/Restore Testing"}:
        return "Major"
    if test_type in {"Accessibility Testing", "Localization Testing", "UX Testing", "Responsive Testing"}:
        return "Moderate"
    return "Major" if priority in {"Critical", "High"} else "Minor"


def functional_case(feature: Feature, feature_index: int, template_index: int, case_index: int) -> dict[str, str]:
    if template_index == 0:
        test_type = select_primary_type(feature.category, feature_index)
        scenario = f"Verify the primary {feature.feature.lower()} workflow completes successfully on {feature.screen}"
        steps = numbered_steps(
            f"Open RESUMIX AI Android and navigate to {feature.screen}.",
            f"Execute the main workflow to {feature.objective} using {feature.sample_data}.",
            "Save, submit, or continue through the dependent flow without leaving the app in an ambiguous state.",
            "Return to the originating surface and confirm the completed action is immediately visible.",
        )
        test_data = feature.sample_data
        expected = f"The app lets the user {feature.objective}, all visible states update correctly, and no crash, freeze, or incorrect navigation occurs."
    elif template_index == 1:
        test_type = "Validation Testing"
        scenario = f"Validate required-field rules and input guardrails for {feature.feature.lower()}"
        steps = numbered_steps(
            f"Navigate to {feature.screen} with the feature ready for manual input.",
            f"Leave mandatory inputs blank or partially completed while attempting to {feature.objective}.",
            "Trigger validation by saving, submitting, or progressing to the next step.",
            "Review inline validation, button state, and any corrective guidance shown to the tester.",
        )
        test_data = feature.invalid_data
        expected = "Mandatory input rules are enforced consistently, the app blocks unsafe progression, and validation feedback is clear, actionable, and visually stable."
    elif template_index == 2:
        test_type = ["Functional Testing", "System Testing"][feature_index % 2]
        scenario = f"Verify {feature.feature.lower()} handles invalid, unavailable, or rejected conditions gracefully"
        steps = numbered_steps(
            f"Reach {feature.screen}.",
            f"Attempt the workflow using invalid or unsupported data: {feature.invalid_data}.",
            "Induce the most likely rejection condition for the module, such as offline mode, cancelation, stale session, or unsupported state.",
            "Observe whether the user can recover without corrupting data or getting stuck.",
        )
        test_data = feature.invalid_data
        expected = f"{feature.feature} fails safely with a user-friendly error, no hidden data corruption, and a clear path to retry or cancel."
    elif template_index == 3:
        test_type = ["Functional Testing", "UX Testing"][feature_index % 2]
        scenario = f"Verify {feature.feature.lower()} supports boundary and real-world data variations"
        steps = numbered_steps(
            f"Open {feature.screen}.",
            f"Use realistic boundary data: {feature.boundary_data}.",
            "Review field behavior, truncation, scrolling, control states, and downstream rendering after save/submit.",
            "Confirm the user experience remains readable and predictable under near-limit input conditions.",
        )
        test_data = feature.boundary_data
        expected = "Boundary inputs are accepted or rejected according to product rules, layouts remain usable, and the saved representation stays readable in downstream screens."
    elif template_index == 4:
        test_type = ["Data Persistence Testing", "Recovery Testing"][feature_index % 2]
        scenario = f"Verify {feature.feature.lower()} persists correctly across app relaunch, reload, and revisit"
        steps = numbered_steps(
            f"Perform {feature.objective} using {feature.sample_data}.",
            "Leave the current screen, return to it, and verify the latest state is still present.",
            "Force close the app or relaunch it from the recent apps tray.",
            "Re-open the same feature and verify the saved state, ordering, and derived UI are intact.",
        )
        test_data = feature.sample_data
        expected = "Saved state persists accurately after navigation, relaunch, or reload; no duplicate, missing, or reordered data appears unexpectedly."
    elif template_index == 5:
        test_type = select_integration_type(feature, feature_index)
        scenario = f"Verify downstream integrations remain correct after using {feature.feature.lower()}"
        steps = numbered_steps(
            f"Open {feature.screen} and complete the feature using {feature.sample_data}.",
            f"Trigger the dependent integration path involving {feature.integration}.",
            "Inspect the receiving screen, service response, or exported artifact for mapping accuracy.",
            "Repeat the same path once more to confirm deterministic behavior and no duplicate side effects.",
        )
        test_data = f"Primary data: {feature.sample_data}; Integration focus: {feature.integration}"
        expected = f"The integration path connected to {feature.integration} receives the correct data, preserves identifiers/titles/order, and surfaces actionable failures if the dependency is unavailable."
    elif template_index == 6:
        test_type = ["Navigation Testing", "UI Testing"][feature_index % 2]
        scenario = f"Verify navigation affordances and UI feedback around {feature.feature.lower()} are intuitive"
        steps = numbered_steps(
            f"Approach {feature.screen} from the nearest realistic entry point in the app.",
            "Use visible buttons, cards, toolbar actions, bottom sheets, and back-navigation controls to complete the flow.",
            "Observe scroll behavior, focus movement, button labels, action ordering, and visual emphasis.",
            "Exit the feature and confirm the user lands in the expected previous context without data loss.",
        )
        test_data = "Mixed normal user interactions including taps, back gestures, and sheet dismissals"
        expected = "Navigation labels and controls are understandable, the user can move in and out of the feature predictably, and the UI remains visually balanced and consistent."
    elif template_index == 7:
        test_type = select_compatibility_type(feature_index)
        scenario = f"Verify {feature.feature.lower()} remains usable across devices, accessibility modes, and localization pressure"
        steps = numbered_steps(
            f"Open {feature.screen} on at least two target device profiles.",
            "Repeat the key task with large font size, screen rotation, dark/light system preference where applicable, and TalkBack or similar assistive mode.",
            "Review long text, truncated labels, contrast, focus order, and tappable area spacing.",
            "Switch to a long-text locale or pseudo-localized copy and inspect whether the feature still reads clearly.",
        )
        test_data = f"{feature.boundary_data}; locale: en-IN and pseudo-long-text; accessibility font scale 1.5x"
        expected = "The feature stays readable and operable across targeted phones, tablets, and accessibility settings, without clipped text, invisible controls, or broken navigation."
    elif template_index == 8:
        test_type = select_interrupt_type(feature_index)
        scenario = f"Verify {feature.feature.lower()} recovers safely from interruptions, resource pressure, and unstable conditions"
        steps = numbered_steps(
            f"Start the workflow on {feature.screen} using {feature.sample_data}.",
            "Interrupt the flow with backgrounding, app switching, notification shade, incoming call simulation, or temporary network loss.",
            "Return to the feature, then continue, retry, or cancel the action.",
            "Observe memory/performance symptoms, unsaved-state handling, and whether the user can complete the task safely.",
        )
        test_data = "Transient network loss, background/foreground cycle, rotation, and low-memory simulation"
        expected = "The feature either resumes gracefully or fails safely with recoverable guidance; no duplicate actions, corrupted data, or prolonged UI stalls occur."
    else:
        test_type = select_regression_type(feature, feature_index)
        scenario = f"Verify {feature.feature.lower()} remains release-ready under regression, security, and production checks"
        steps = numbered_steps(
            f"Open {feature.screen} on a production-like build.",
            "Run the highest-risk scenario for the feature, including entitlement, identity, file integrity, or privacy-sensitive behavior where relevant.",
            "Compare the observed behavior with the expected production policy, release checklist, and core workflow output.",
            "Confirm the feature can be signed off without blocking release, or capture the precise deviation.",
        )
        test_data = f"Production-like configuration with {feature.sample_data}"
        expected = "The feature passes its regression and release-readiness checks, respects privacy/security expectations, and remains stable for Play Store production deployment."

    priority = priority_for(feature, test_type, template_index)
    severity = severity_for(feature, test_type, template_index, priority)

    return {
        "Test Case ID": f"RMX-F-{case_index:04d}",
        "Test Type": test_type,
        "Module Name": feature.module,
        "Feature Name": feature.feature,
        "Test Scenario": scenario,
        "Preconditions": build_preconditions(feature, template_index),
        "Test Steps": steps,
        "Test Data": test_data,
        "Expected Result": expected,
        "Actual Result": "Pending execution",
        "Status": "Not Executed",
        "Priority": priority,
        "Severity": severity,
        "Environment": choose_environment(feature.category, test_type),
        "Device/Platform": choose_device(feature_index, template_index, test_type),
        "Build Version": BUILD_VERSION,
        "Remarks/Comments": f"Generated enterprise manual test case covering {TEMPLATE_LABELS[template_index].lower()} for {feature.screen}.",
    }


def exploratory_case(feature: Feature, feature_index: int, charter_index: int, case_index: int) -> dict[str, str]:
    if charter_index == 0:
        scenario = f"Explore the first-time and golden-path journey for {feature.feature.lower()}"
        steps = numbered_steps(
            f"Start from the earliest realistic entry point that leads to {feature.screen}.",
            f"Use {feature.sample_data} to pursue the most valuable user journey and note every decision point.",
            "Observe cognitive friction, hidden dependencies, loading states, and confidence signals for a first-time user.",
            "Record discoveries around value clarity, missing guidance, and any blocker or high-friction transitions.",
        )
        data = feature.sample_data
        remarks = "Charter focus: first-use experience, confidence, discoverability, and business-value flow."
    elif charter_index == 1:
        scenario = f"Explore invalid inputs, strange content, and misuse around {feature.feature.lower()}"
        steps = numbered_steps(
            f"Open {feature.screen} and intentionally vary the inputs beyond normal expectations.",
            f"Try malformed, conflicting, empty, pasted, duplicated, and semi-valid data such as {feature.invalid_data} and {feature.boundary_data}.",
            "Look for misleading errors, silent failures, hidden truncation, duplicate writes, or unexpectedly accepted data.",
            "Document the most surprising system behavior and the exact reproduction path.",
        )
        data = f"Invalid data: {feature.invalid_data}; Boundary data: {feature.boundary_data}"
        remarks = "Charter focus: edge-case discovery, validation gaps, and data-quality risks."
    elif charter_index == 2:
        scenario = f"Explore interruption, poor connectivity, and recovery behavior for {feature.feature.lower()}"
        steps = numbered_steps(
            f"Begin the feature flow on {feature.screen} using {feature.sample_data}.",
            "Introduce backgrounding, network drops, device rotation, app kill/relaunch, or notification interruptions at different stages.",
            "Attempt to recover, retry, or cancel from every interrupted state.",
            "Capture any stale state, duplicated action, missing content, or user confusion that appears after recovery.",
        )
        data = "Offline/online transitions, background-foreground cycles, notification interruptions, force close"
        remarks = "Charter focus: resilience, recovery quality, and data-safety observations."
    else:
        scenario = f"Explore responsive, accessibility, and release-readiness heuristics for {feature.feature.lower()}"
        steps = numbered_steps(
            f"Run the feature on {feature.screen} across at least two devices and one accessibility configuration.",
            "Vary font scale, long text, landscape/portrait, TalkBack, and a pseudo-localized or long-string data set.",
            "Observe reachability, focus order, contrast, clipped text, responsive breaks, and policy-sensitive copy or disclosures.",
            "Log findings that would affect production readiness, supportability, or Play Store review confidence.",
        )
        data = f"Devices: {DEVICE_POOL[feature_index % len(DEVICE_POOL)]}, {DEVICE_POOL[(feature_index + 4) % len(DEVICE_POOL)]}; Locale stress + font scale 1.5x"
        remarks = "Charter focus: adaptive UI, accessibility heuristics, localization stress, and production readiness."

    priority = BASE_PRIORITY[feature.category]
    severity = "Major" if priority in {"Critical", "High"} else "Moderate"

    return {
        "Test Case ID": f"RMX-E-{case_index:04d}",
        "Test Type": "Exploratory Testing",
        "Module Name": feature.module,
        "Feature Name": feature.feature,
        "Test Scenario": scenario,
        "Preconditions": build_preconditions(feature, charter_index),
        "Test Steps": steps,
        "Test Data": data,
        "Expected Result": "The session should reveal no blocker, critical data loss, or unacceptable UX/policy issue; meaningful observations and defects can be captured with reproducible notes.",
        "Actual Result": "Pending exploration",
        "Status": "Not Executed",
        "Priority": priority,
        "Severity": severity,
        "Environment": choose_environment(feature.category, "Exploratory Testing"),
        "Device/Platform": choose_device(feature_index, charter_index, "Cross-device Testing"),
        "Build Version": BUILD_VERSION,
        "Remarks/Comments": remarks,
    }


def build_functional_cases() -> list[dict[str, str]]:
    cases: list[dict[str, str]] = []
    case_index = 1
    for feature_index, feature in enumerate(FEATURES):
        for template_index in range(10):
            cases.append(functional_case(feature, feature_index, template_index, case_index))
            case_index += 1
    if len(cases) != 500:
        raise ValueError(f"Expected 500 functional/manual cases, generated {len(cases)}")
    return cases


def build_exploratory_cases() -> list[dict[str, str]]:
    cases: list[dict[str, str]] = []
    case_index = 1
    for feature_index, feature in enumerate(FEATURES):
        for charter_index in range(4):
            cases.append(exploratory_case(feature, feature_index, charter_index, case_index))
            case_index += 1
    if len(cases) != 200:
        raise ValueError(f"Expected 200 exploratory cases, generated {len(cases)}")
    return cases


def assert_quality(cases: list[dict[str, str]], label: str) -> None:
    unique_keys = {
        (case["Module Name"], case["Feature Name"], case["Test Scenario"])
        for case in cases
    }
    if len(unique_keys) != len(cases):
        raise ValueError(f"Duplicate scenarios detected in {label}")


def build_value_palette(values: list[str], palette: list[str]) -> dict[str, str]:
    unique_values = list(dict.fromkeys(values))
    return {
        value: palette[index % len(palette)]
        for index, value in enumerate(unique_values)
    }


def autofit_columns(ws) -> None:
    custom_widths = {
        "A": 14,
        "B": 24,
        "C": 22,
        "D": 30,
        "E": 42,
        "F": 36,
        "G": 52,
        "H": 34,
        "I": 44,
        "J": 24,
        "K": 16,
        "L": 12,
        "M": 12,
        "N": 28,
        "O": 28,
        "P": 24,
        "Q": 34,
    }
    for column_letter, width in custom_widths.items():
        ws.column_dimensions[column_letter].width = width


def style_data_sheet(ws, rows: list[dict[str, str]], table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    module_colors = build_value_palette([row["Module Name"] for row in rows], MODULE_PALETTE)
    test_type_colors = build_value_palette([row["Test Type"] for row in rows], TEST_TYPE_PALETTE)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_index in range(2, ws.max_row + 1):
        for column_index in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.alignment = wrap_alignment
            cell.border = border

        ws.cell(row=row_index, column=2).fill = PatternFill("solid", fgColor=test_type_colors[ws.cell(row=row_index, column=2).value])
        ws.cell(row=row_index, column=3).fill = PatternFill("solid", fgColor=module_colors[ws.cell(row=row_index, column=3).value])
        ws.cell(row=row_index, column=11).fill = PatternFill("solid", fgColor=STATUS_COLORS[ws.cell(row=row_index, column=11).value])
        ws.cell(row=row_index, column=12).fill = PatternFill("solid", fgColor=PRIORITY_COLORS[ws.cell(row=row_index, column=12).value])
        ws.cell(row=row_index, column=13).fill = PatternFill("solid", fgColor=SEVERITY_COLORS[ws.cell(row=row_index, column=13).value])

    ws.freeze_panes = "A2"
    autofit_columns(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_simple_sheet(ws, table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    widths = {"A": 18, "B": 42, "C": 14}
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width

    table = Table(displayName=table_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_rows(ws, rows: list[dict[str, str]]) -> None:
    ws.append(MANDATORY_COLUMNS)
    for row in rows:
        ws.append([row[column] for column in MANDATORY_COLUMNS])


def build_summary_sheet(ws, functional_cases: list[dict[str, str]], exploratory_cases: list[dict[str, str]]) -> None:
    title_fill = PatternFill("solid", fgColor="0B5394")
    accent_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = "RESUMIX AI Android Enterprise QA Test Suite"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    summary_lines = [
        ("Generated On", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Workbook Output", str(OUTPUT_FILE)),
        ("Functional/Manual Cases", str(len(functional_cases))),
        ("Exploratory Cases", str(len(exploratory_cases))),
        ("Total Cases", str(len(functional_cases) + len(exploratory_cases))),
        ("Modules Covered", str(len({case['Module Name'] for case in functional_cases + exploratory_cases}))),
        ("Distinct Test Types", str(len({case['Test Type'] for case in functional_cases + exploratory_cases}))),
        ("Build Baseline", BUILD_VERSION),
    ]

    for offset, (label, value) in enumerate(summary_lines, start=3):
        ws[f"A{offset}"] = label
        ws[f"B{offset}"] = value
        ws[f"A{offset}"].font = Font(bold=True)
        ws[f"A{offset}"].fill = accent_fill
        ws[f"A{offset}"].border = border
        ws[f"B{offset}"].border = border

    ws["D3"] = "Production-ready coverage focus"
    ws["D3"].font = Font(bold=True)
    ws["D3"].fill = accent_fill
    ws["D3"].border = border
    production_notes = [
        "Authentication via OTP and Google sign-in",
        "Resume CRUD, editor flows, custom sections, and reorder persistence",
        "Preview, PDF export, file validation, and freemium rules",
        "AI generation, rewrite, tailoring, ATS, roast, and LinkedIn import",
        "Cloud backup/restore, subscription billing, settings/privacy, and data deletion",
        "Responsive, accessibility, localization, compatibility, performance, and Play Store readiness",
    ]
    for offset, note in enumerate(production_notes, start=4):
        ws[f"D{offset}"] = note
        ws[f"D{offset}"].border = border
        ws[f"D{offset}"].alignment = Alignment(wrap_text=True, vertical="top")

    module_counts = Counter(case["Module Name"] for case in functional_cases + exploratory_cases)
    type_counts = Counter(case["Test Type"] for case in functional_cases + exploratory_cases)

    start_row = 12
    ws[f"A{start_row}"] = "Module Coverage"
    ws[f"A{start_row}"].font = Font(bold=True)
    ws[f"A{start_row}"].fill = accent_fill
    ws[f"A{start_row}"].border = border
    ws[f"A{start_row + 1}"] = "Module Name"
    ws[f"B{start_row + 1}"] = "Case Count"
    for cell in (ws[f"A{start_row + 1}"], ws[f"B{start_row + 1}"]):
        cell.font = Font(bold=True)
        cell.fill = title_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = border
    row = start_row + 2
    for module_name, count in module_counts.most_common():
        ws[f"A{row}"] = module_name
        ws[f"B{row}"] = count
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border
        row += 1

    ws[f"D{start_row}"] = "Test Type Coverage"
    ws[f"D{start_row}"].font = Font(bold=True)
    ws[f"D{start_row}"].fill = accent_fill
    ws[f"D{start_row}"].border = border
    ws[f"D{start_row + 1}"] = "Test Type"
    ws[f"E{start_row + 1}"] = "Case Count"
    for cell in (ws[f"D{start_row + 1}"], ws[f"E{start_row + 1}"]):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = title_fill
        cell.border = border
    row = start_row + 2
    for test_type, count in type_counts.most_common():
        ws[f"D{row}"] = test_type
        ws[f"E{row}"] = count
        ws[f"D{row}"].border = border
        ws[f"E{row}"].border = border
        row += 1

    ws["A33"] = "Notes"
    ws["A33"].font = Font(bold=True)
    ws["A33"].fill = accent_fill
    ws["A33"].border = border
    notes = [
        "All rows are pre-populated as Not Executed / Pending execution so QA can record actual results during runs.",
        "Color coding is applied to module, test type, priority, severity, and status columns for audit and presentation use.",
        "Workbook includes 500 functional/manual cases and 200 exploratory charters with full Android production-readiness scope.",
    ]
    for offset, note in enumerate(notes, start=34):
        ws[f"A{offset}"] = note
        ws[f"A{offset}"].alignment = Alignment(wrap_text=True, vertical="top")

    for column_letter, width in {"A": 26, "B": 28, "D": 38, "E": 16}.items():
        ws.column_dimensions[column_letter].width = width


def main() -> None:
    functional_cases = build_functional_cases()
    exploratory_cases = build_exploratory_cases()
    assert_quality(functional_cases, "functional suite")
    assert_quality(exploratory_cases, "exploratory suite")

    actual_test_types = {case["Test Type"] for case in functional_cases + exploratory_cases}
    missing_types = [test_type for test_type in ALL_REQUIRED_TEST_TYPES if test_type not in actual_test_types]
    if missing_types:
        raise ValueError(f"Missing required test types: {missing_types}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    build_summary_sheet(summary_ws, functional_cases, exploratory_cases)

    functional_ws = workbook.create_sheet("Functional_Manual_500")
    add_rows(functional_ws, functional_cases)
    style_data_sheet(functional_ws, functional_cases, "FunctionalManualSuite")

    exploratory_ws = workbook.create_sheet("Exploratory_200")
    add_rows(exploratory_ws, exploratory_cases)
    style_data_sheet(exploratory_ws, exploratory_cases, "ExploratorySuite")

    coverage_ws = workbook.create_sheet("Coverage_Matrix")
    coverage_ws.append(["Dimension", "Value", "Case Count"])
    combined_cases = functional_cases + exploratory_cases
    for module_name, count in Counter(case["Module Name"] for case in combined_cases).most_common():
        coverage_ws.append(["Module", module_name, count])
    for test_type, count in Counter(case["Test Type"] for case in combined_cases).most_common():
        coverage_ws.append(["Test Type", test_type, count])
    style_simple_sheet(coverage_ws, "CoverageMatrix")

    workbook.save(OUTPUT_FILE)
    print(f"Generated workbook: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()