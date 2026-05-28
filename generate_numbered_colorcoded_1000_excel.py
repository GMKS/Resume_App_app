from __future__ import annotations

import colorsys
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WORKSPACE_ROOT = Path(__file__).resolve().parent
SOURCE_CSV = WORKSPACE_ROOT / "resume_app_test_suite_1000_requested_columns.csv"
OUTPUT_XLSX = WORKSPACE_ROOT / "resume_app_test_suite_1000_numbered_colorcoded.xlsx"

REQUIRED_COLUMNS = [
    "Test Case ID",
    "Test Scenario",
    "Test Case Title",
    "Module Name",
    "Requirement ID",
    "Preconditions",
    "Test Steps",
    "Test Data",
    "Expected Result",
    "Actual Result",
]


def convert_to_numbered_text(text: str | None) -> str:
    if not text or not text.strip():
        return ""

    normalized = " ".join(text.replace("\r", " ").replace("\n", " ").split())
    if not normalized:
        return ""

    parts: list[str] = []
    numbered_chunks = []
    index = 0
    while index < len(normalized):
        if normalized[index].isdigit():
            number_end = index
            while number_end < len(normalized) and normalized[number_end].isdigit():
                number_end += 1
            if number_end < len(normalized) - 1 and normalized[number_end:number_end + 2] == ". ":
                chunk_start = number_end + 2
                next_index = chunk_start
                while next_index < len(normalized):
                    if normalized[next_index].isdigit():
                        lookahead = next_index
                        while lookahead < len(normalized) and normalized[lookahead].isdigit():
                            lookahead += 1
                        if lookahead < len(normalized) - 1 and normalized[lookahead:lookahead + 2] == ". ":
                            break
                    next_index += 1
                numbered_chunks.append(normalized[chunk_start:next_index].strip())
                index = next_index
                continue
        index += 1

    if numbered_chunks:
        parts = [chunk for chunk in numbered_chunks if chunk]
    else:
        sentence = []
        for token in normalized.split(" "):
            sentence.append(token)
            if token.endswith((".", "!", "?")):
                parts.append(" ".join(sentence).strip())
                sentence = []
        if sentence:
            parts.append(" ".join(sentence).strip())

    return "\n".join(f"{i}. {part}" for i, part in enumerate(parts, start=1) if part)


def hsl_fill(index: int, total: int) -> PatternFill:
    hue = (index / max(total, 1)) % 1.0
    red, green, blue = colorsys.hls_to_rgb(hue, 0.83, 0.45)
    rgb = f"{int(red * 255):02X}{int(green * 255):02X}{int(blue * 255):02X}"
    return PatternFill(fill_type="solid", start_color=rgb, end_color=rgb)


def main() -> None:
    if not SOURCE_CSV.exists():
        raise FileNotFoundError(f"Source CSV not found: {SOURCE_CSV}")

    with SOURCE_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)

    if not rows:
        raise ValueError("Source CSV is empty.")

    missing = [column for column in REQUIRED_COLUMNS if column not in rows[0]]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    module_order: list[str] = []
    seen_modules: set[str] = set()
    for row in rows:
        module_name = row["Module Name"]
        if module_name not in seen_modules:
            seen_modules.add(module_name)
            module_order.append(module_name)

    module_fills = {
        module_name: hsl_fill(index, len(module_order))
        for index, module_name in enumerate(module_order)
    }
    requirement_by_module = {
        module_name: next(
            row["Requirement ID"] for row in rows if row["Module Name"] == module_name
        )
        for module_name in module_order
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "1000 Test Cases"
    legend_sheet = workbook.create_sheet("Module Legend")

    header_fill = PatternFill(fill_type="solid", start_color="1F2937", end_color="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    centered_alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for column_index, header in enumerate(REQUIRED_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = thin_border

    for row_index, row in enumerate(rows, start=2):
        values = [
            row["Test Case ID"],
            row["Test Scenario"],
            row["Test Case Title"],
            row["Module Name"],
            row["Requirement ID"],
            convert_to_numbered_text(row["Preconditions"]),
            convert_to_numbered_text(row["Test Steps"]),
            row["Test Data"],
            convert_to_numbered_text(row["Expected Result"]),
            row["Actual Result"],
        ]

        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if column_index == 4:
                cell.fill = module_fills[row["Module Name"]]
                cell.font = Font(bold=True, color="111827")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{len(rows) + 1}"
    sheet.sheet_view.showGridLines = True
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 40
    sheet.column_dimensions["D"].width = 34
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 46
    sheet.column_dimensions["G"].width = 54
    sheet.column_dimensions["H"].width = 28
    sheet.column_dimensions["I"].width = 46
    sheet.column_dimensions["J"].width = 22

    for row_index in range(2, len(rows) + 2):
        sheet.row_dimensions[row_index].height = 68

    legend_headers = ["Module Name", "Requirement ID", "Color Key"]
    for column_index, header in enumerate(legend_headers, start=1):
        cell = legend_sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = thin_border

    for row_index, module_name in enumerate(module_order, start=2):
        legend_sheet.cell(row=row_index, column=1, value=module_name)
        legend_sheet.cell(row=row_index, column=2, value=requirement_by_module[module_name])
        swatch = legend_sheet.cell(row=row_index, column=3, value=" ")
        swatch.fill = module_fills[module_name]
        for column_index in range(1, 4):
            legend_sheet.cell(row=row_index, column=column_index).border = thin_border
            legend_sheet.cell(row=row_index, column=column_index).alignment = centered_alignment if column_index == 3 else cell_alignment

    legend_sheet.freeze_panes = "A2"
    legend_sheet.auto_filter.ref = f"A1:C{len(module_order) + 1}"
    legend_sheet.column_dimensions["A"].width = 38
    legend_sheet.column_dimensions["B"].width = 16
    legend_sheet.column_dimensions["C"].width = 14

    workbook.save(OUTPUT_XLSX)
    print(f"Generated Excel workbook: {OUTPUT_XLSX}")
    print(f"Rows exported: {len(rows)}")
    print(f"Modules color-coded: {len(module_order)}")


if __name__ == "__main__":
    main()