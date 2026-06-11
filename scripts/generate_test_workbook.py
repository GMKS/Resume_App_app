"""
Resume Builder App - Professional Test Case Workbook Generator
Generates a color-coded Excel workbook with 500+ test cases
across 18 testing types.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'build', 'reports',
                           'ResumeBuilder_TestCaseWorkbook.xlsx')

# ── Column definitions ───────────────────────────────────────────────────────
COLUMNS = [
    ('Test Case ID',        18),
    ('Module',              20),
    ('Scenario',            28),
    ('Test Case Description', 48),
    ('Priority',            12),
    ('Preconditions',       38),
    ('Test Data',           38),
    ('Test Steps',          60),
    ('Expected Result',     48),
    ('Actual Result',       30),
    ('Status',              14),
    ('Defect ID',           14),
    ('Comments',            30),
    ('Automation Candidate', 20),
]

# ── Sheet tab config: (display name, tab hex colour, prefix) ─────────────────
SHEETS = [
    ('Functional',          'FF4472C4', 'FN'),
    ('End-to-End E2E',      'FFED7D31', 'E2E'),
    ('Smoke',               'FFA9D18E', 'SMK'),
    ('Sanity',              'FF70AD47', 'SAN'),
    ('Regression',          'FFFFC000', 'REG'),
    ('UI-UX',               'FFFF0000', 'UI'),
    ('API',                 'FF7030A0', 'API'),
    ('Integration',         'FF0070C0', 'INT'),
    ('System',              'FF00B0F0', 'SYS'),
    ('Security',            'FFFF0000', 'SEC'),
    ('Performance',         'FFFF7F00', 'PER'),
    ('Compatibility',       'FF548235', 'CMP'),
    ('Accessibility',       'FF833C00', 'ACC'),
    ('Negative',            'FFBF8F00', 'NEG'),
    ('Boundary',            'FF403151', 'BND'),
    ('Subscription-Billing','FFD6E4BC', 'SUB'),
    ('Backup-Sync',         'FFB4C6E7', 'BAK'),
    ('Cross-Platform',      'FFF4B942', 'CPL'),
]

# ── Style helpers ─────────────────────────────────────────────────────────────
def header_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color='FFD0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def make_header_style(hex_color):
    # Determine if white or dark font gives better contrast
    r = int(hex_color[2:4], 16)
    g = int(hex_color[4:6], 16)
    b = int(hex_color[6:8], 16)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    font_color = 'FFFFFFFF' if luminance < 0.6 else 'FF1F1F1F'
    return (
        PatternFill('solid', fgColor=hex_color),
        Font(bold=True, color=font_color, size=10),
    )

ALT_ROW_FILL = PatternFill('solid', fgColor='FFF2F2F2')
WHITE_FILL   = PatternFill('solid', fgColor='FFFFFFFF')

# ── Status dropdown colours ───────────────────────────────────────────────────
STATUS_COLOURS = {
    'Not Run':  'FFD9D9D9',
    'Pass':     'FF92D050',
    'Fail':     'FFFF4B4B',
    'Blocked':  'FFFFC000',
    'N/A':      'FFD9D9D9',
}

PRIORITY_COLOURS = {
    'Critical': 'FFFF4B4B',
    'High':     'FFFFC000',
    'Medium':   'FF00B0F0',
    'Low':      'FFA9D18E',
}

def priority_fill(p):
    return PatternFill('solid', fgColor=PRIORITY_COLOURS.get(p, 'FFFFFFFF'))


# ═══════════════════════════════════════════════════════════════════════════════
# TEST DATA
# Each row: (Module, Scenario, Description, Priority, Preconditions,
#            Test Data, Test Steps, Expected Result, Automation Candidate)
# ═══════════════════════════════════════════════════════════════════════════════

# ── FUNCTIONAL ───────────────────────────────────────────────────────────────
FUNCTIONAL = [
    ('Authentication','OTP Login – Valid','User can log in using a valid mobile number and OTP','Critical','App installed; network available; valid SIM-based mobile number','Mobile: +91-9876543210; OTP: received from Twilio','1. Open app\n2. Tap "Login with Phone"\n3. Enter valid mobile\n4. Tap "Send OTP"\n5. Enter received OTP\n6. Tap "Verify"','User is authenticated and redirected to Home/Dashboard','','Not Run','','','Yes'),
    ('Authentication','OTP Login – Invalid OTP','System rejects incorrect OTP','High','App installed; OTP request sent','Mobile: +91-9876543210; Wrong OTP: 000000','1. Open app\n2. Enter valid mobile\n3. Send OTP\n4. Enter wrong OTP\n5. Tap "Verify"','Error message "Invalid OTP" shown; user stays on verify screen','','Not Run','','','Yes'),
    ('Authentication','Google Sign-In','User can sign in via Google account','Critical','Google account available; internet connected','Gmail: testuser@gmail.com','1. Tap "Continue with Google"\n2. Select Google account\n3. Grant permissions','User signed in; profile info populated from Google account','','Not Run','','','Yes'),
    ('Authentication','Facebook Sign-In','User can sign in via Facebook','High','Facebook app or browser available','FB account: testuser@example.com','1. Tap "Continue with Facebook"\n2. Log in to FB\n3. Authorise app','User authenticated; lands on Dashboard','','Not Run','','','No'),
    ('Authentication','Guest / Skip Login','User can use app without signing in','Medium','App installed','N/A','1. Tap "Skip" or "Continue as Guest"','App opens with limited features; no cloud sync enabled','','Not Run','','','Yes'),
    ('Authentication','Logout','User can log out from Settings','High','User logged in','Logged in session','1. Open Settings\n2. Scroll to Account\n3. Tap "Logout"\n4. Confirm','User logged out; app returns to Login screen','','Not Run','','','Yes'),
    ('Authentication','Session Persistence','Session persists after app restart','High','User logged in','Active session','1. Log in\n2. Close app completely\n3. Reopen app','User is still logged in; lands on Home','','Not Run','','','Yes'),
    ('Resume Creation','Create New Resume','User creates a new resume from scratch','Critical','User logged in; on Home tab','Name: "John Doe", Title: "Software Engineer"','1. Tap "+" / New Resume\n2. Enter name and target role\n3. Tap "Create"','New blank resume created and editor opens','','Not Run','','','Yes'),
    ('Resume Creation','Duplicate Resume','User can duplicate an existing resume','Medium','At least one resume exists','Existing resume: "Resume_A"','1. Long-press or tap menu on resume card\n2. Select "Duplicate"','Exact copy created with "(Copy)" suffix','','Not Run','','','Yes'),
    ('Resume Creation','Delete Resume','User can delete a resume','High','At least one resume exists','Resume: "Resume_A"','1. Swipe left on resume card or tap menu\n2. Select "Delete"\n3. Confirm','Resume deleted; no longer appears in list','','Not Run','','','Yes'),
    ('Resume Editor','Personal Info – Save','All personal info fields saved correctly','Critical','Resume created','Name: "Jane Smith"; Email: "jane@test.com"; Phone: "+1-555-0199"','1. Open editor\n2. Tap "Personal Info"\n3. Fill all fields\n4. Tap "Save"','Fields saved; visible in preview','','Not Run','','','Yes'),
    ('Resume Editor','Add Work Experience','User adds work experience entry','Critical','Resume in edit mode','Company: "Acme Corp"; Role: "Dev"; Dates: "Jan 2020 – Dec 2022"','1. Tap "Experience"\n2. Tap "Add"\n3. Fill all fields\n4. Save','Entry saved and rendered in resume preview','','Not Run','','','Yes'),
    ('Resume Editor','Edit Work Experience','Existing experience entry can be edited','High','One experience entry exists','Changed role: "Senior Dev"','1. Tap existing experience\n2. Edit Role field\n3. Save','Updated text reflects in preview','','Not Run','','','Yes'),
    ('Resume Editor','Delete Work Experience','Experience entry removed successfully','High','One or more experience entries','Entry: "Acme Corp"','1. Swipe or tap delete on experience entry\n2. Confirm','Entry removed; preview updates instantly','','Not Run','','','Yes'),
    ('Resume Editor','Add Education','User adds an education entry','Critical','Resume in edit mode','Institution: "MIT"; Degree: "BSc CS"; Year: 2018','1. Tap "Education"\n2. Add entry\n3. Fill fields\n4. Save','Entry saved and appears in preview','','Not Run','','','Yes'),
    ('Resume Editor','Add Skills','User adds skills with proficiency','High','Resume in edit mode','Skill: "Python"; Level: "Expert"','1. Tap "Skills"\n2. Tap "Add Skill"\n3. Enter name and level\n4. Save','Skill visible in preview with correct level','','Not Run','','','Yes'),
    ('Resume Editor','Add Summary','User adds a professional summary','High','Resume in edit mode','Summary: 3-sentence paragraph','1. Tap "Summary"\n2. Type summary\n3. Save','Summary saved and displayed in preview','','Not Run','','','Yes'),
    ('Resume Editor','Add Certifications','User adds certification entry','Medium','Resume in edit mode','Cert: "AWS SAA"; Issuer: "Amazon"; Date: "2023"','1. Tap "Certifications"\n2. Add entry\n3. Fill fields\n4. Save','Certification appears in resume','','Not Run','','','Yes'),
    ('Resume Editor','Add Projects','User adds a project entry','Medium','Resume in edit mode','Project: "eCommerce App"; Stack: "Flutter, Firebase"','1. Tap "Projects"\n2. Add entry\n3. Fill all fields\n4. Save','Project appears in resume preview','','Not Run','','','Yes'),
    ('Resume Editor','Add Languages','User adds spoken languages','Low','Resume in edit mode','Language: "Spanish"; Level: "Intermediate"','1. Tap "Languages"\n2. Add\n3. Fill\n4. Save','Language entry visible in preview','','Not Run','','','Yes'),
    ('Resume Editor','Add Custom Section','User adds a custom section','Medium','Resume in edit mode','Section: "Volunteer Work"','1. Tap "Add Section"\n2. Enter section title\n3. Add content\n4. Save','Custom section appears in resume','','Not Run','','','Yes'),
    ('Resume Editor','Profile Photo Upload','User attaches profile photo','Medium','Camera/gallery permission granted','Image: JPG < 5 MB','1. Tap profile photo placeholder\n2. Choose Gallery\n3. Select image','Photo appears in resume header in preview','','Not Run','','','Yes'),
    ('Template Selection','Change Template','User changes resume template','Critical','Resume with content exists','Template: "Corporate Navy"','1. Open resume editor\n2. Tap template/paintbrush icon\n3. Select "Corporate Navy"\n4. Confirm','Resume re-renders with new template; content preserved','','Not Run','','','Yes'),
    ('Template Selection','Preview All Templates','User can browse all 28+ templates','High','Resume exists','N/A','1. Open template picker\n2. Scroll through all options','All templates visible with thumbnail previews','','Not Run','','','Yes'),
    ('Template Selection','Colour Customisation','User changes accent colour','Medium','Template selected','Colour: #E74C3C','1. Open template settings\n2. Tap colour picker\n3. Select custom colour','Resume preview updates with new accent colour','','Not Run','','','Yes'),
    ('Preview & Export','PDF Preview','Resume renders correctly in preview','Critical','Resume with content; template selected','Full resume data','1. Tap "Preview"\n2. Scroll through all sections','PDF preview renders all sections without clipping or data loss','','Not Run','','','Yes'),
    ('Preview & Export','Download PDF','User downloads PDF to device','Critical','Preview screen open','Full resume data','1. Tap "Download PDF"\n2. Confirm save location','PDF saved to device; file is valid and readable','','Not Run','','','Yes'),
    ('Preview & Export','Share PDF','User shares PDF via external app','High','Preview screen open','Full resume data','1. Tap "Share"\n2. Select WhatsApp or Gmail','Share sheet opens; PDF attachment correct','','Not Run','','','Yes'),
    ('Preview & Export','Print Resume','User prints resume','Medium','Printer or Google Cloud Print configured','Full resume data','1. Tap "Print"\n2. Select printer','Print dialog opens; page layout correct','','Not Run','','','Yes'),
    ('Home / Dashboard','Resume List Display','All resumes shown on Home tab','High','Multiple resumes created','3 resumes','1. Navigate to Home tab\n2. Observe resume cards','All resumes listed with name, template badge, last-modified date','','Not Run','','','Yes'),
    ('Home / Dashboard','Resume Search','User can search resumes by name','Medium','3+ resumes exist','Search: "Engineer"','1. Tap search bar\n2. Type "Engineer"','Only matching resumes shown','','Not Run','','','Yes'),
    ('Settings','Theme Toggle – Dark','App switches to dark mode','Medium','App running in light mode','N/A','1. Open Settings\n2. Tap "Theme"\n3. Select "Dark Mode"','Entire UI switches to dark theme','','Not Run','','','Yes'),
    ('Settings','Theme Toggle – System','Theme follows device setting','Low','App running','Device theme: Dark','1. Set theme to "System Default"\n2. Toggle device dark mode','App theme mirrors device setting','','Not Run','','','Yes'),
    ('Settings','App Version Display','Correct version shown in About','Low','App installed','Version: 1.0.2','1. Open Settings\n2. Tap "About"','Version string matches pubspec version','','Not Run','','','Yes'),
    ('Career Tools','Job Tracker – Add Job','User adds a job application','High','Career Tools tab open','Company: "Google"; Status: "Applied"','1. Tap "Job Tracker"\n2. Tap "+" \n3. Fill details\n4. Save','Job card appears in tracker list','','Not Run','','','Yes'),
    ('Career Tools','Cover Letter Generator','User generates a cover letter','High','AI configured; premium plan','Job: "Flutter Developer at Meta"','1. Tap "AI Cover Letter"\n2. Enter job details\n3. Generate','Cover letter text generated and editable','','Not Run','','','Yes'),
    ('Career Tools','Skill Analyzer','Skill recommendations shown by role','Medium','Premium plan active','Role: "Data Scientist"','1. Tap "Skill Analyzer"\n2. Enter role\n3. Analyse','Role-specific skill list displayed with gap analysis','','Not Run','','','Yes'),
    ('Career Tools','Interview Prep','Q&A generated for target role','Medium','AI configured','Role: "Product Manager"','1. Tap "Interview Prep"\n2. Enter role\n3. Generate','Relevant interview questions and model answers shown','','Not Run','','','Yes'),
    ('Onboarding','First-Launch Onboarding','New users see onboarding flow','High','Fresh install; no account','N/A','1. Install app\n2. Open for first time','Onboarding slides displayed; can skip or proceed','','Not Run','','','Yes'),
    ('AI – Resume Generator','Generate Resume from Scratch','AI generates a complete resume','Critical','AI service configured; premium plan','Prompt: "Senior Flutter Developer, 5 yrs exp"','1. Tap "AI Resume Generator"\n2. Enter role and context\n3. Tap "Generate"','Complete resume sections generated in <30 s','','Not Run','','','Yes'),
    ('AI – Content Enhancer','Enhance Experience Bullet','AI rewrites experience bullet professionally','High','AI configured; free tier has limit','Bullet: "worked on stuff"','1. Tap "AI Content Enhancer"\n2. Paste bullet\n3. Enhance','Professionally rewritten bullet returned','','Not Run','','','Yes'),
    ('AI – Job Tailor','Match Score','Resume scored against job description','High','AI configured','Resume + JD pasted','1. Tap "Resume Match Analyzer"\n2. Paste JD\n3. Analyse','Match % and missing skills shown','','Not Run','','','Yes'),
    ('AI – Resume Rewrite','Full Rewrite','AI rewrites entire resume','High','Premium plan; AI configured','Existing resume','1. Tap "AI Resume Rewrite"\n2. Confirm\n3. Generate','New resume with stronger language returned','','Not Run','','','Yes'),
    ('AI – Bullet Generator','Generate Bullets','Role-based bullets generated','High','AI configured','Role: "Backend Engineer"','1. Tap "AI Bullet Generator"\n2. Enter role and context\n3. Generate','5-7 strong ATS bullets returned','','Not Run','','','Yes'),
    ('LinkedIn Import','Import LinkedIn Profile','Resume auto-built from LinkedIn','High','LinkedIn account; premium plan','LinkedIn URL or pasted profile text','1. Tap "LinkedIn Import"\n2. Paste profile text\n3. Import','Resume sections populated from LinkedIn data','','Not Run','','','Yes'),
]

# ── END-TO-END (E2E) ──────────────────────────────────────────────────────────
E2E = [
    ('Full Journey','New User → Create → Export PDF','New user signs up, creates resume, exports PDF','Critical','Fresh install; internet; Google account','Google: newuser@test.com; Full resume data','1. Install app\n2. Sign in with Google\n3. Skip onboarding\n4. Tap "+ New Resume"\n5. Fill all sections\n6. Select template\n7. Preview\n8. Download PDF','PDF file created on device with correct content; no data loss','','Not Run','','','Yes'),
    ('Full Journey','Login → Edit Existing Resume → Share','Returning user edits resume and shares','High','Account with saved resume','Existing resume with partial data','1. Open app (session restored)\n2. Open resume\n3. Edit experience\n4. Preview\n5. Share via Gmail','Share intent fires with attached PDF','','Not Run','','','Yes'),
    ('Full Journey','AI Resume Generation → Template → Export','AI-generated resume customised and exported','High','AI service available; premium','Prompt: "Marketing Manager 7 yrs"','1. Tap AI Resume Generator\n2. Enter prompt\n3. Select generated content\n4. Change template\n5. Preview\n6. Download PDF','PDF exported with AI content and selected template styling','','Not Run','','','Yes'),
    ('Full Journey','Job Tracker Full Workflow','Add → update → delete job application','High','Career Tools tab accessible','Job: "Apple - iOS Engineer"','1. Add job\n2. Change status to "Interview Scheduled"\n3. Add note\n4. Change to "Offer Received"\n5. Delete entry','All status transitions persist; deletion clears entry','','Not Run','','','Yes'),
    ('Full Journey','Subscription → Premium Feature Unlock','Purchase plan, unlock AI cover letter','Critical','Razorpay configured; test card','Test card: 4111 1111 1111 1111','1. Open Settings → Upgrade\n2. Select Monthly plan\n3. Complete payment\n4. Return to AI Cover Letter','Cover letter feature unlocked; "Premium" badge shown','','Not Run','','','Yes'),
    ('Full Journey','Backup → New Device Restore','Backup on Device A; restore on Device B','High','Supabase sync configured; 2 devices','Sync Code: "e2e-test-001"','1. Device A: set sync code\n2. Backup\n3. Device B: set same code\n4. Restore','All resumes and job tracker data appear on Device B','','Not Run','','','Yes'),
    ('Full Journey','LinkedIn Import → Edit → Export','Import profile, customise, export','High','LinkedIn text; AI available','LinkedIn plain text profile','1. LinkedIn Import\n2. Review imported sections\n3. Edit summary\n4. Change template\n5. Export PDF','PDF reflects imported + edited content correctly','','Not Run','','','Yes'),
    ('Full Journey','Phone OTP → Resume → Logout → Re-login','Full auth cycle with phone','High','Valid mobile number; Twilio','Mobile: +91-9876543210','1. Login via OTP\n2. Create resume\n3. Logout\n4. Re-login with same number','Resume data still present after re-login','','Not Run','','','Yes'),
    ('Full Journey','ATS Analysis → Optimise → Re-score','Run ATS score, apply suggestions, re-score','Medium','Resume with content; AI available','JD: Software Engineer at Google','1. Open Resume Match Analyzer\n2. Run initial score\n3. Apply top 3 suggestions\n4. Re-run analysis','Second score is higher; suggestions applied correctly','','Not Run','','','Yes'),
    ('Full Journey','Cover Letter → Job Tracker Link','Generate cover letter for tracked job','Medium','Premium plan; job in tracker','Job: "Spotify – Product Lead"','1. Open job tracker entry\n2. Tap "Generate Cover Letter"\n3. Review\n4. Save to device','Cover letter generated and linked to job entry','','Not Run','','','Yes'),
    ('Full Journey','Onboarding → Template Browse → Resume Create','New user completes onboarding and creates first resume','High','Fresh install','N/A','1. Onboarding\n2. Browse templates\n3. Select template\n4. Create resume\n5. Fill Personal Info\n6. Save','Resume with chosen template saved successfully','','Not Run','','','Yes'),
    ('Full Journey','Resume Versioning','Multiple saves create version history','Medium','Resume with 3+ edits','Edits at different times','1. Open resume\n2. Make edit A, save\n3. Make edit B, save\n4. Open version history','At least 2 prior versions listed; restore works','','Not Run','','','Yes'),
    ('Full Journey','Multi-Resume Switch','User switches between two open resumes','Medium','2 resumes created','Resume A and Resume B','1. Open Resume A, note content\n2. Back to list\n3. Open Resume B\n4. Back to list\n5. Reopen Resume A','Each resume shows its own distinct content without cross-contamination','','Not Run','','','Yes'),
    ('Full Journey','Delete All Data → Fresh Start','User deletes all data and starts over','High','Multiple resumes and jobs saved','N/A','1. Settings → Delete All Data\n2. Confirm\n3. Create new resume','All prior data gone; new resume saved cleanly','','Not Run','','','Yes'),
    ('Full Journey','Portfolio Generation Flow','User generates portfolio from resume','Medium','Premium; resume with content','Full resume','1. Open resume\n2. Tap Portfolio icon\n3. Generate\n4. Preview portfolio URL','Portfolio page generated with correct sections','','Not Run','','','No'),
]

# ── SMOKE ─────────────────────────────────────────────────────────────────────
SMOKE = [
    ('App Launch','App Opens Without Crash','App launches to login/home screen','Critical','App installed on device','N/A','1. Tap app icon','Splash screen shown then login or home screen loads in <4 s','','Not Run','','','Yes'),
    ('Authentication','OTP Screen Reachable','OTP login screen loads','Critical','App installed','N/A','1. Tap "Login with Phone"','Phone number input screen displayed','','Not Run','','','Yes'),
    ('Authentication','Google Sign-In Initiates','Google auth sheet appears','Critical','Google Play Services available','N/A','1. Tap "Continue with Google"','Google account picker appears','','Not Run','','','Yes'),
    ('Home','Resume List Loads','Home tab shows resume list (or empty state)','Critical','User logged in','N/A','1. Navigate to Home tab','Resume list or empty state visible; no crash','','Not Run','','','Yes'),
    ('Resume Creation','Create New Resume Flow Starts','New resume wizard/form opens','Critical','On Home tab','N/A','1. Tap "+" button','New resume creation flow opens','','Not Run','','','Yes'),
    ('Resume Editor','Editor Sections Visible','All editor sections accessible','Critical','Resume created','N/A','1. Open resume\n2. Scroll section list','Personal Info, Experience, Education, Skills, Summary visible','','Not Run','','','Yes'),
    ('Template Selection','Template Picker Opens','Template selection screen loads','Critical','Resume in editor','N/A','1. Tap template/paintbrush icon','Template grid/list loads without crash','','Not Run','','','Yes'),
    ('Preview','Preview Screen Loads','PDF preview renders','Critical','Resume with content','N/A','1. Tap "Preview"','Preview renders in <5 s; no blank page','','Not Run','','','Yes'),
    ('Export','Download PDF Works','PDF file created on device','Critical','Preview open','N/A','1. Tap "Download PDF"','File saved; success toast shown','','Not Run','','','Yes'),
    ('Settings','Settings Screen Opens','Settings page loads','Critical','Logged in','N/A','1. Tap Settings icon','Settings screen with all sections loads','','Not Run','','','Yes'),
    ('AI','AI Assistant Screen Loads','AI tool list renders','Critical','AI configured or not','N/A','1. Tap AI Assistant nav item','AI tool cards visible','','Not Run','','','Yes'),
    ('Career Tools','Career Tools Tab Loads','Career tools menu loads','Critical','Logged in','N/A','1. Tap Career Tools tab','All career tool cards displayed','','Not Run','','','Yes'),
    ('Subscription','Upgrade Screen Loads','Subscription/upgrade screen accessible','Critical','Logged in (free plan)','N/A','1. Tap locked feature or Upgrade button','Subscription plans screen loads','','Not Run','','','Yes'),
    ('Backup','Backup & Sync Sheet Opens','Backup bottom sheet renders','High','Logged in; premium or feature enabled','N/A','1. Settings → Backup & Sync','Backup sheet opens with backup/restore buttons','','Not Run','','','Yes'),
    ('Navigation','All Bottom Nav Tabs Work','All tabs reachable without crash','Critical','Logged in','N/A','1. Tap each bottom nav tab in sequence','Each tab loads its screen without error','','Not Run','','','Yes'),
    ('Connectivity','Offline State Handled','App does not crash when offline','High','WiFi off; mobile data off','Airplane mode','1. Enable airplane mode\n2. Open app','App shows offline message or cached content; no crash','','Not Run','','','Yes'),
    ('Crash-Free','Background → Foreground','App resumes cleanly from background','Critical','App running','N/A','1. Open app\n2. Home button\n3. Re-open','App returns to exact previous state','','Not Run','','','Yes'),
    ('Crash-Free','Rapid Screen Switching','Quick tab navigation does not crash','High','Logged in','N/A','1. Rapidly tap all tabs 10× each','No crash; no navigation stack corruption','','Not Run','','','Yes'),
    ('Crash-Free','Low Memory Simulation','App survives low memory pressure','High','Developer options available','N/A','1. Enable "Don\'t keep activities"\n2. Navigate app','App restores state correctly; no crash','','Not Run','','','Yes'),
    ('Crash-Free','Rotate Screen','Landscape rotation handled','Medium','Portrait mode default','N/A','1. Open resume editor\n2. Rotate to landscape','UI adapts without crash or content loss','','Not Run','','','Yes'),
]

# ── SANITY ────────────────────────────────────────────────────────────────────
SANITY = [
    ('Build Verification','Latest Build Launches','New build opens without crash','Critical','APK/AAB installed','Version: latest','1. Install new build\n2. Open app','App launches; correct version shown in About','','Not Run','','','Yes'),
    ('Authentication','Login Works Post-Build','OTP login successful on new build','Critical','New build installed','Valid mobile','1. Login via OTP','Auth succeeds on latest build','','Not Run','','','Yes'),
    ('Resume','Create + Save Verified','Resume create and save work on new build','Critical','Logged in','Name: "Test User"','1. Create resume\n2. Add name\n3. Save','Resume appears in list','','Not Run','','','Yes'),
    ('Resume','PDF Export Works','PDF generated without error on new build','Critical','Resume with content','N/A','1. Open resume\n2. Preview\n3. Download','PDF file saved to device','','Not Run','','','Yes'),
    ('AI','AI Feature Responds','AI returns a response (smoke pass)','High','AI configured','Prompt: "Software Engineer"','1. Tap AI Content Enhancer\n2. Enter text\n3. Enhance','Response returned in <15 s','','Not Run','','','Yes'),
    ('Subscription','Upgrade Plan Screen Accessible','Plan screen loads on latest build','High','Logged in; free plan','N/A','1. Tap any locked feature','Subscription plans shown','','Not Run','','','Yes'),
    ('Settings','Theme Toggle Works','Dark/light theme switching works','Medium','Logged in','N/A','1. Toggle dark mode','UI switches theme instantly','','Not Run','','','Yes'),
    ('Career Tools','Job Tracker Accessible','Job tracker loads on latest build','Medium','Career Tools tab','N/A','1. Open Job Tracker','Job list or empty state shown','','Not Run','','','Yes'),
    ('Backup','Backup Button Functional','Backup process initiates','High','Premium plan; internet','Sync code set','1. Backup & Sync → Backup to Cloud','Progress indicator shown; success message returned','','Not Run','','','Yes'),
    ('Navigation','Back Navigation Works','Back button returns to previous screen','High','Multiple screens open','N/A','1. Navigate deep (Resume → Editor → Experience)\n2. Press back','Each back step returns to correct prior screen','','Not Run','','','Yes'),
    ('Template','Template Change Persists','Template selection survives app restart','High','Resume with template selected','N/A','1. Select template\n2. Close and reopen app','Same template active on reopen','','Not Run','','','Yes'),
    ('Data Persistence','Resume Data Persists','Resume content unchanged after restart','Critical','Resume with full data','N/A','1. Fill resume\n2. Kill app\n3. Reopen','All fields intact','','Not Run','','','Yes'),
]

# ── REGRESSION ────────────────────────────────────────────────────────────────
REGRESSION = [
    ('Authentication','OTP Resend Cooldown Respected','Cannot spam resend before cooldown','High','OTP request sent','Mobile: +91-9876543210','1. Request OTP\n2. Immediately tap Resend\n3. Observe','Resend disabled for 60 s; countdown shown','','Not Run','','','Yes'),
    ('Resume Editor','Auto-Save on Section Exit','Edits auto-saved when leaving section','High','Resume open in editor','New experience text','1. Open Experience\n2. Type entry\n3. Tap back without explicit save','Data retained when re-entering section','','Not Run','','','Yes'),
    ('Resume Editor','Special Characters in Fields','Special chars handled correctly in PDF','High','Resume editor','Name: "O\'Brien & Smith <Sr.>"','1. Enter special chars in name field\n2. Preview PDF','Name renders exactly in PDF; no encoding errors','','Not Run','','','Yes'),
    ('Resume Editor','Long Text Truncation','Very long text doesn\'t break layout','Medium','Resume editor','Summary: 2000 characters','1. Enter very long summary\n2. Preview PDF','Text wraps correctly; no overflow clipping in PDF','','Not Run','','','Yes'),
    ('Template','Colour Accent Persists After Edit','Accent colour retained after editor changes','Medium','Custom colour applied','Colour: #8E44AD','1. Set custom colour\n2. Edit experience section\n3. Preview','Same custom colour in preview','','Not Run','','','Yes'),
    ('Template','All 28 Templates Render','Every template renders without crash','High','Resume with full data','N/A','1. For each template, select and preview','All templates render PDF successfully','','Not Run','','','Yes'),
    ('AI','AI Feature Rate Limit','Free tier limit message shown at threshold','High','Free plan; AI used 4 times (limit test)','4 prior AI calls','1. Open AI Content Enhancer\n2. Use 5th time','"Limit reached" or upgrade prompt shown','','Not Run','','','Yes'),
    ('Subscription','Expired Plan Reverts Features','Premium features locked after plan expires','Critical','Premium plan that has expired','Expired plan account','1. Open app with expired plan\n2. Try AI Cover Letter','Premium feature locked; upgrade prompt shown','','Not Run','','','Yes'),
    ('Backup','Sync Code Change Isolates Data','New sync code creates isolated cloud space','High','Prior backup with code "test-001"','New code: "test-002"','1. Change sync code to "test-002"\n2. Backup\n3. Change to "test-001"\n4. Restore','Restoring old code brings back original data; no mix','','Not Run','','','Yes'),
    ('Data','Delete All Does Not Delete Account','Delete Data clears resumes but keeps account','High','Logged in with account and resumes','N/A','1. Settings → Delete All Data\n2. Confirm\n3. Check account status','Resumes deleted; account still logged in','','Not Run','','','Yes'),
    ('Navigation','Deep-Link Navigation','Deep link opens correct screen','Medium','App installed','Deep link: /ai-resume-generator','1. Trigger deep link\n2. Observe destination','AI Resume Generator screen opens directly','','Not Run','','','Yes'),
    ('PDF','Duplicate Name Handling in PDF','Two identical names don\'t conflict in export','Medium','Two resumes with same name','Name: "My Resume"','1. Create two resumes named "My Resume"\n2. Export both as PDF','Both PDFs created; files distinguished by timestamp','','Not Run','','','Yes'),
    ('Settings','Rate App Opens Play Store','Rate App link opens correct Play page','Low','Internet available','N/A','1. Settings → Rate App\n2. Tap','Play Store opens on the app\'s page','','Not Run','','','No'),
    ('Profile Photo','Photo Removed Clears Field','Removing photo updates preview','Medium','Photo set on resume','Existing photo','1. Tap profile photo\n2. Remove photo','Photo removed from preview and PDF','','Not Run','','','Yes'),
    ('Resume','Resume Name Editing','Resume title can be renamed','High','Existing resume','Old name: "Draft 1"','1. Tap rename on resume card\n2. Enter "Final Resume"\n3. Save','Resume card shows "Final Resume"','','Not Run','','','Yes'),
    ('Career Tools','Job Status Transition','Job status changes tracked correctly','High','Job in tracker','Job: "Apple - Engineer"','1. Change status from "Applied" to "Rejected"','Status chip updates; history recorded','','Not Run','','','Yes'),
    ('AI','LinkedIn Import Preserves Formatting','Imported text sections intact','Medium','Valid LinkedIn text input','3 experience entries','1. Paste LinkedIn profile text\n2. Import','All 3 experience entries mapped to correct fields','','Not Run','','','Yes'),
    ('Accessibility','Screen Reader Not Broken by Update','TalkBack still navigable post-build','Medium','TalkBack enabled','N/A','1. Enable TalkBack\n2. Navigate Home, Editor, Settings','No silent focusable areas; all labels read aloud','','Not Run','','','No'),
    ('Performance','Cold Start Under 4 s Regression','Build regression does not slow start','High','Clean device state','N/A','1. Kill app\n2. Tap icon\n3. Measure time to first frame','App interactive within 4 seconds','','Not Run','','','Yes'),
    ('Security','JWT Token Not Logged','No auth token appears in debug logs after fix','Critical','Debug build','N/A','1. Login\n2. Open Logcat/console\n3. Search "token"','No plaintext token in logs','','Not Run','','','Yes'),
]

# ── UI/UX ─────────────────────────────────────────────────────────────────────
UI_UX = [
    ('Layout','Home Screen Layout','Cards aligned, margins consistent','High','Logged in; 3 resumes','N/A','1. Open Home tab\n2. Observe resume card layout','Cards evenly spaced; no overflow; correct shadows','','Not Run','','','No'),
    ('Layout','Editor Section Headers','Section titles bold and readable','Medium','Resume in editor','N/A','1. Scroll through editor sections','All section headers visible in correct font weight','','Not Run','','','No'),
    ('Typography','Font Consistency','Uniform font family across all screens','Medium','All screens accessible','N/A','1. Navigate through app\n2. Note font usage','Consistent type face (Google Fonts) throughout app','','Not Run','','','No'),
    ('Colour','Dark Mode Contrast','All text readable in dark mode','High','Dark mode enabled','N/A','1. Enable dark mode\n2. Navigate app','No white-on-white or black-on-black text anywhere','','Not Run','','','No'),
    ('Colour','Light Mode Contrast','All text readable in light mode','High','Light mode','N/A','1. Navigate all screens in light mode','Sufficient contrast ratio (WCAG AA) on all text','','Not Run','','','No'),
    ('Animations','Screen Transitions Smooth','Page transitions are 60 fps','Medium','App running','N/A','1. Navigate between 5 screens rapidly','No jank or dropped frames visible','','Not Run','','','No'),
    ('Animations','Card Entry Animation','Resume cards animate on load','Low','Home tab','N/A','1. Open Home tab fresh','Cards fade/slide in smoothly','','Not Run','','','No'),
    ('Touch Targets','All Buttons Tappable','Minimum 48×48 dp touch targets','High','App running','N/A','1. Inspect all interactive elements','No buttons smaller than 48 dp','','Not Run','','','No'),
    ('Touch Targets','FAB Placement','Floating Action Button does not obscure content','Medium','Home tab','N/A','1. Scroll resume list with FAB visible','FAB does not hide last resume card entirely','','Not Run','','','No'),
    ('Error States','Empty Resume List State','Friendly empty state shown','High','No resumes created','N/A','1. Delete all resumes\n2. Open Home tab','Illustration + "Create your first resume" CTA visible','','Not Run','','','No'),
    ('Error States','Network Error State','Friendly error for AI when offline','High','Airplane mode','N/A','1. Offline\n2. Tap AI tool','Error message with retry option; no raw exception shown','','Not Run','','','No'),
    ('Feedback','Loading Indicators','Progress shown for long operations','High','Normal network','N/A','1. Trigger backup, AI call, PDF export','Spinner or progress bar visible during operation','','Not Run','','','No'),
    ('Feedback','Success Toasts','Positive feedback for save/export','Medium','N/A','N/A','1. Save resume\n2. Download PDF','Green toast with message visible 2-3 s','','Not Run','','','No'),
    ('Feedback','Error Toasts','Clear error messages for failures','High','Network error simulated','N/A','1. Trigger AI call offline','Red/orange toast with readable error','','Not Run','','','No'),
    ('Keyboard','Keyboard Avoidance','Content not hidden by keyboard','High','Editor screen, phone','N/A','1. Tap text field in editor\n2. Keyboard opens','Text field visible above keyboard; no overlap','','Not Run','','','No'),
    ('Keyboard','Done Button Dismisses Keyboard','Done/Return closes keyboard','Medium','Text field focused','N/A','1. Tap text field\n2. Tap done on keyboard','Keyboard dismisses; field retains text','','Not Run','','','No'),
    ('Scroll','Long Resume Scrollable','Editor scrolls to all sections','High','Resume with 10+ sections','N/A','1. Open full resume\n2. Scroll to bottom','All sections reachable without dead zone','','Not Run','','','No'),
    ('Responsive','Tablet Layout','UI adapts to 10-inch tablet','Medium','Tablet device or emulator (1280×800)','N/A','1. Open app on tablet\n2. Navigate main screens','Layout uses wider space; no stretched single-column cards','','Not Run','','','No'),
    ('Branding','App Icon Correct','Correct icon shown on home screen','Low','App installed','N/A','1. View app icon on launcher','Branded icon shown; no default Flutter icon','','Not Run','','','No'),
    ('Branding','Splash Screen','Splash screen shows brand logo','Low','Cold start','N/A','1. Kill app\n2. Launch','Branded splash before main app','','Not Run','','','No'),
    ('Bottom Sheet','Bottom Sheets Draggable','Draggable sheets expand/collapse','Medium','Any bottom sheet open','N/A','1. Open backup sheet\n2. Drag up and down','Sheet expands and collapses smoothly','','Not Run','','','No'),
    ('Modal','Dialogs Dismissible','Dialogs can be dismissed by tapping outside','Medium','Dialog open','N/A','1. Open confirm dialog\n2. Tap outside','Dialog dismisses (or shows cancel confirmation)','','Not Run','','','No'),
    ('RTL','LTR-only layout does not break Arabic locale','No UI breakage on Arabic locale','Low','Device language: Arabic','N/A','1. Set device language to Arabic\n2. Open app','App displays without overflow; text input works','','Not Run','','','No'),
]

# ── API ────────────────────────────────────────────────────────────────────────
API = [
    ('Twilio OTP','Send OTP – Valid','Valid mobile receives OTP','Critical','Twilio configured; valid mobile','Mobile: +91-9876543210','1. POST OTP send endpoint\n2. Valid body','HTTP 200; OTP delivered to phone within 60 s','','Not Run','','','Yes'),
    ('Twilio OTP','Send OTP – Invalid Number','Invalid number returns error','High','Twilio configured','Mobile: +99-0000000000','1. POST with invalid number','HTTP 400 or Twilio error code; friendly error in UI','','Not Run','','','Yes'),
    ('Twilio OTP','Verify OTP – Valid','Correct OTP accepted','Critical','OTP sent to valid number','Correct OTP code','1. POST verify endpoint with correct OTP','HTTP 200; auth token returned','','Not Run','','','Yes'),
    ('Twilio OTP','Verify OTP – Expired','Expired OTP rejected','High','OTP older than 10 min','Stale OTP code','1. POST verify with stale OTP','HTTP 401 or "OTP expired" error','','Not Run','','','Yes'),
    ('Twilio OTP','Rate Limit OTP Send','Repeated OTP sends rejected after threshold','High','Twilio rate-limit configured','5+ rapid requests','1. POST OTP send 6 times rapidly','HTTP 429 or rate-limit error after threshold','','Not Run','','','Yes'),
    ('GROQ AI','Generate AI Content – Valid','Valid prompt returns AI response','Critical','GROQ_API_KEY configured','Prompt: "Improve: worked on backend"','1. POST /ai/enhance with valid body','HTTP 200; improved text in response body','','Not Run','','','Yes'),
    ('GROQ AI','Generate AI Content – Empty Prompt','Empty prompt returns validation error','High','AI configured','Prompt: ""','1. POST with empty prompt field','HTTP 400; "prompt is required" error','','Not Run','','','Yes'),
    ('GROQ AI','AI Timeout Handling','Slow AI response handled gracefully','Medium','AI configured; simulated slow endpoint','Prompt: long input','1. Trigger AI call\n2. Simulate 30 s delay','Timeout error shown in UI; no app freeze','','Not Run','','','Yes'),
    ('Supabase Sync','Backup Resumes – Valid','Resume data backed up to Supabase','Critical','Supabase configured; sync code set','Resume JSON payload','1. POST backup with valid resume payload','HTTP 200 or 201; data in Supabase table','','Not Run','','','Yes'),
    ('Supabase Sync','Restore Resumes – Valid','Resume data fetched from Supabase','Critical','Backup exists in Supabase','Sync code: "e2e-test-001"','1. GET restore with valid sync code','HTTP 200; resume JSON matches backed-up data','','Not Run','','','Yes'),
    ('Supabase Sync','Backup – Unauthorized','No sync code returns error','High','Supabase configured','No sync code','1. POST backup without sync code','HTTP 401 or RLS error; no data written','','Not Run','','','Yes'),
    ('Supabase Sync','Restore – Wrong Code','Wrong sync code returns empty','High','Supabase configured','Sync code: "wrong-code-xyz"','1. GET restore with wrong code','HTTP 200 with empty array; no data leak','','Not Run','','','Yes'),
    ('Razorpay','Initiate Payment – Valid','Checkout opened with correct amount','Critical','Razorpay key configured; test mode','Plan: Monthly ₹199; test card','1. Select plan\n2. Tap "Subscribe"\n3. Observe Razorpay sheet','Razorpay checkout opens with correct amount and currency','','Not Run','','','Yes'),
    ('Razorpay','Payment Success Callback','Premium unlocked after payment success','Critical','Payment completed','Razorpay test card success','1. Complete payment\n2. Observe app state','Premium plan activated; features unlocked','','Not Run','','','Yes'),
    ('Razorpay','Payment Failure Callback','Graceful error on payment failure','High','Razorpay test failure card','Test card: fail scenario','1. Use failure test card\n2. Attempt payment','Payment failed message shown; plan stays free','','Not Run','','','Yes'),
    ('Google Play Billing','Product List Fetched','Play SKUs fetched correctly','High','Play Billing configured; test track','N/A','1. Open subscription screen\n2. Observe product pricing','Correct prices from Play fetched and displayed','','Not Run','','','Yes'),
    ('Google Play Billing','Purchase Acknowledged','Purchase acknowledged to prevent refund','Critical','Play Billing; test account','Premium SKU purchased','1. Complete Play purchase\n2. Check acknowledgement','Purchase acknowledged within 3 days; no auto-refund','','Not Run','','','Yes'),
    ('HTTP Security','All API Calls Use HTTPS','No plain HTTP requests','Critical','Proxy tool (Charles/mitmproxy)','N/A','1. Intercept all network traffic during normal use','All requests over HTTPS; no HTTP URLs','','Not Run','','','Yes'),
    ('HTTP Security','No API Key in Request Body','GROQ key not visible in logs/body','Critical','Proxy tool','N/A','1. Intercept AI API call\n2. Inspect request body','API key not present in body or URL params','','Not Run','','','Yes'),
    ('HTTP Headers','Auth Header Present','Authenticated requests send auth token','High','Logged-in session','N/A','1. Trigger backup API\n2. Inspect headers','Authorization header present with valid token','','Not Run','','','Yes'),
]

# ── INTEGRATION ───────────────────────────────────────────────────────────────
INTEGRATION = [
    ('Auth + Storage','Login Syncs User Profile to Storage','User profile saved locally after login','High','Google Sign-In','Google account','1. Sign in with Google\n2. Check local profile store','Profile name/email stored locally','','Not Run','','','Yes'),
    ('Resume + PDF','Resume Data → PDF Generator','Resume model correctly passed to PDF engine','Critical','Resume with all sections filled','Full resume JSON','1. Open resume\n2. Preview PDF','All sections present in PDF','','Not Run','','','Yes'),
    ('AI + Resume','AI-Generated Text Inserted into Resume','AI bullet inserted into experience field','High','AI configured; resume open','Role: "Backend Engineer"','1. Open experience section\n2. Use AI Bullet Generator\n3. Insert suggested bullet','Bullet appears in experience field','','Not Run','','','Yes'),
    ('Subscription + AI','Premium Plan Unlocks AI Cover Letter','Subscription state controls AI cover letter access','Critical','Premium purchased','N/A','1. Purchase premium\n2. Open AI Cover Letter','Feature accessible without upgrade prompt','','Not Run','','','Yes'),
    ('Subscription + Free Plan','Free Plan Blocks Premium AI Tools','Free plan user cannot access premium AI','High','Free plan account','N/A','1. Tap AI Resume Rewrite (premium)','Upgrade prompt shown; feature blocked','','Not Run','','','Yes'),
    ('Supabase + Hive','Backup Serialises Hive Objects','Hive resume model serialises for Supabase','High','Resume saved in Hive; Supabase configured','N/A','1. Backup resumes\n2. Inspect Supabase table','Resume JSON matches Hive model structure','','Not Run','','','Yes'),
    ('Supabase + Hive','Restore Deserialises to Hive','Supabase data correctly loads into Hive','High','Backup in Supabase','N/A','1. Restore data\n2. Open resume list','Restored resumes open correctly in editor','','Not Run','','','Yes'),
    ('Riverpod + UI','Provider State Reflects UI Changes','Riverpod state changes update UI instantly','High','App running','N/A','1. Add experience entry\n2. Observe real-time preview','Preview updates without manual refresh','','Not Run','','','Yes'),
    ('GoRouter + Deep Link','Deep Link Navigates Correctly','GoRouter routes deep links properly','Medium','App installed','Deep link: /settings','1. Trigger /settings deep link','Settings screen opens directly','','Not Run','','','Yes'),
    ('Job Tracker + Cover Letter','Job Entry Used in Cover Letter Prompt','Job details auto-populate cover letter form','Medium','Job in tracker; AI available','Job: "Meta – Product Manager"','1. Open job entry\n2. Tap "Generate Cover Letter"','Job title and company pre-filled in cover letter form','','Not Run','','','Yes'),
    ('Firebase Auth + Social','Firebase Auth State Synced with App State','Firebase auth changes trigger app state update','High','Firebase configured','N/A','1. Sign in with Google\n2. Revoke in Firebase console\n3. Return to app','App detects de-auth; returns to login screen','','Not Run','','','Yes'),
    ('In-App Purchase + Subscription Model','IAP Purchase Updates SubscriptionModel','Completed IAP propagates to Riverpod subscription state','Critical','Play Billing; test account','Premium SKU','1. Complete IAP\n2. Observe subscription provider','isPremium = true; features unlocked','','Not Run','','','Yes'),
    ('Resume Quality + Editor','Quality Score Updates on Edit','Quality score recalculates on content change','Medium','Resume with quality score','N/A','1. Open resume\n2. Add experience\n3. Check quality score','Score changes after adding experience','','Not Run','','','Yes'),
    ('Translation + Resume','Translated Resume Exports as PDF','Translated text renders in PDF','Medium','LibreTranslate configured; resume in English','Target: Spanish','1. Translate resume to Spanish\n2. Export PDF','Spanish text in exported PDF','','Not Run','','','No'),
    ('Version Service + Build','Build Number Increments on Each Build','Build number monotonically increases','High','Two sequential builds','N/A','1. Build once, note version\n2. Build again','Second build has higher versionCode','','Not Run','','','Yes'),
]

# ── SYSTEM ────────────────────────────────────────────────────────────────────
SYSTEM = [
    ('Device','Android 10 Compatibility','App runs on Android 10 (API 29)','High','Android 10 device/emulator','N/A','1. Install APK on Android 10\n2. Run core flows','All critical flows work on Android 10','','Not Run','','','No'),
    ('Device','Android 13 Compatibility','App runs on Android 13 (API 33)','High','Android 13 device','N/A','1. Install on Android 13\n2. Run core flows','All critical flows work on Android 13','','Not Run','','','No'),
    ('Device','Android 14 Compatibility','App runs on Android 14 (API 34)','High','Android 14 device','N/A','1. Install on Android 14\n2. Run core flows','All features functional on Android 14','','Not Run','','','No'),
    ('Storage','Low Storage Warning Handled','App handles <100 MB free storage','Medium','Device with <100 MB free','N/A','1. Fill device storage\n2. Try to export PDF','Friendly error shown; no crash','','Not Run','','','No'),
    ('Storage','Large Resume Dataset','App handles 50+ resumes without slowdown','Medium','50 resumes in local storage','50 test resumes','1. Load 50 resumes\n2. Navigate Home tab','Home tab loads in <3 s; no ANR','','Not Run','','','No'),
    ('Memory','Memory Leak Check','No significant memory growth over 30-min session','High','Profiling tool (Android Studio)','N/A','1. Use app for 30 min across all features\n2. Monitor heap','Heap does not grow unboundedly','','Not Run','','','No'),
    ('Battery','Background Battery Usage','App does not drain battery in background','Medium','Battery stats tool','N/A','1. Run app for 10 min\n2. Background 30 min\n3. Check battery stats','No wake locks held; <1% battery in background','','Not Run','','','No'),
    ('Permissions','Camera Permission Request','Camera permission requested only when needed','High','Fresh install','N/A','1. Tap profile photo\n2. Select camera','Permission dialog appears first time only','','Not Run','','','No'),
    ('Permissions','Storage Permission Request','Storage permission for PDF export','High','Android <10 device','N/A','1. Tap Download PDF','Storage permission requested; PDF saved after grant','','Not Run','','','No'),
    ('Permissions','Permission Denied Handling','App handles denied permissions gracefully','High','Permission denied','N/A','1. Deny camera permission\n2. Tap photo upload','Friendly message shown; no crash','','Not Run','','','No'),
    ('Network','Slow 3G Network','App usable on slow network','Medium','Network throttled to 3G','N/A','1. Throttle to 3G\n2. Perform AI call and backup','Operations complete or time out gracefully; no crash','','Not Run','','','No'),
    ('Network','Network Switch Mid-Operation','WiFi→Mobile switch during backup','Medium','WiFi and mobile data available','N/A','1. Start backup on WiFi\n2. Disable WiFi mid-operation','Operation retries or shows error; no data corruption','','Not Run','','','No'),
    ('File System','PDF Saved to Correct Location','PDF stored in Downloads folder','High','Android 10+','N/A','1. Export PDF\n2. Check Downloads folder','PDF present in Downloads with correct filename','','Not Run','','','No'),
    ('Concurrency','Simultaneous AI Calls','Multiple AI calls do not corrupt state','Medium','AI configured','N/A','1. Trigger two AI calls quickly\n2. Observe responses','Each call returns its own correct response; no swap','','Not Run','','','No'),
    ('Startup','Cold Start Performance','Cold start under 4 s on mid-range device','High','Mid-range Android device (Snapdragon 680)','N/A','1. Kill app\n2. Launch\n3. Measure','App interactive in <4 s','','Not Run','','','No'),
    ('Startup','Warm Start Performance','Warm start under 2 s','Medium','App previously running','N/A','1. Background app\n2. Bring to foreground\n3. Measure','App interactive in <2 s','','Not Run','','','No'),
]

# ── SECURITY ──────────────────────────────────────────────────────────────────
SECURITY = [
    ('Data Storage','Credentials Not in SharedPreferences','No plaintext credentials in SharedPreferences','Critical','Root access or ADB pull','N/A','1. Login\n2. Pull SharedPreferences XML via ADB\n3. Inspect','No passwords or tokens in plaintext in SharedPreferences','','Not Run','','','No'),
    ('Data Storage','Secrets in Flutter Secure Storage','API keys in FlutterSecureStorage (Keystore)','Critical','Root device or ADB shell','N/A','1. Set API key in app\n2. ADB pull secure storage','Encrypted blob; no plaintext keys visible','','Not Run','','','No'),
    ('Transport','TLS Certificate Pinning','Rejects connections with invalid cert','High','Proxy with custom cert','mitmproxy cert','1. Route traffic through proxy with untrusted cert\n2. Attempt AI/backup call','Connection refused; TLS error shown, not intercepted','','Not Run','','','No'),
    ('Transport','HTTPS Only','No cleartext HTTP allowed','Critical','Network inspector','N/A','1. Monitor all network calls\n2. Check protocols','All calls over TLS 1.2+; no HTTP','','Not Run','','','No'),
    ('Auth','OTP Brute Force Prevention','Account locked after 5 wrong OTPs','High','Twilio rate limiting enabled','Wrong OTPs: 000001–000005','1. Enter wrong OTP 5 times','Account locked or rate limited; no further attempts accepted','','Not Run','','','No'),
    ('Auth','Session Token Expiry','Expired token triggers re-login','High','Session > 24 hrs old','N/A','1. Let session expire\n2. Attempt API call','401 response; app redirects to login','','Not Run','','','No'),
    ('Auth','Token Not in URL','Auth token never appears in URL','Critical','Network log','N/A','1. Inspect all network requests after login','No token in query params or URL path','','Not Run','','','No'),
    ('Input Validation','SQL Injection via Resume Fields','SQL injection in name field is harmless','Critical','Resume editor','Value: \'; DROP TABLE users; --','1. Enter SQL injection string in name field\n2. Save\n3. Export PDF','String stored and displayed literally; no DB error','','Not Run','','','No'),
    ('Input Validation','XSS in Text Fields','HTML/JS in fields does not execute','High','Resume editor','Value: <script>alert(1)</script>','1. Enter script tag in summary\n2. Preview PDF','Script rendered as literal text; no execution','','Not Run','','','No'),
    ('Input Validation','Path Traversal in File Name','Cannot write to arbitrary path via filename','High','Export flow','Filename: ../../etc/passwd.pdf','1. Attempt export with traversal filename','File saved in app-scoped directory only; traversal blocked','','Not Run','','','No'),
    ('Encryption','Local Database Encrypted','Hive boxes encrypted at rest','High','Root device; Hive inspection','N/A','1. Inspect Hive box files\n2. Attempt to read raw','Data not readable without decryption key','','Not Run','','','No'),
    ('Privacy','No PII in Analytics/Logs','No personal data leaked to logs','High','Debug build; Logcat','N/A','1. Login with real name/email\n2. Inspect Logcat for email/name','No email, name, or phone in Logcat output','','Not Run','','','No'),
    ('OWASP M1','Improper Platform Usage – Permissions','App requests only necessary permissions','High','AndroidManifest.xml review','N/A','1. Inspect permissions in manifest','Only INTERNET, CAMERA, WRITE_EXTERNAL declared; no unnecessary perms','','Not Run','','','No'),
    ('OWASP M2','Insecure Data Storage','No sensitive data in plaintext files','Critical','ADB pull of app data dir','N/A','1. ADB pull /data/data/com.resumebuilder.app\n2. Grep for "password"','No passwords or keys in plaintext files','','Not Run','','','No'),
    ('OWASP M5','Insufficient Cryptography','Encryption uses strong algorithms','High','Code review + ProGuard output','N/A','1. Review crypto usage in codebase','AES-256 or equivalent used; no MD5/SHA1 for secrets','','Not Run','','','No'),
    ('OWASP M9','Reverse Engineering Protection','Obfuscation applied in release build','High','Decompile release APK','apktool / jadx','1. Decompile release APK\n2. Inspect class names','Class/method names obfuscated; source not recoverable','','Not Run','','','No'),
    ('Payment','Payment Data Not Stored Locally','Card details never saved to device','Critical','Root device; Razorpay flow','Test card data','1. Complete payment\n2. Inspect local storage','No card number, CVV, or expiry in local storage','','Not Run','','','No'),
    ('Backup','Sync Code Not Exposed in Logs','Sync code not printed in debug output','High','Debug build; Logcat','N/A','1. Set sync code "secret-123"\n2. Backup\n3. Search Logcat','Sync code not logged in plain text','','Not Run','','','No'),
]

# ── PERFORMANCE ───────────────────────────────────────────────────────────────
PERFORMANCE = [
    ('Startup','Cold Start Time – High End','Cold start <2 s on flagship device','High','Pixel 7 / Galaxy S23','N/A','1. Kill app\n2. Launch\n3. Measure via Perfetto','Time to first frame < 2 s','','Not Run','','','Yes'),
    ('Startup','Cold Start Time – Mid Range','Cold start <4 s on mid-range device','High','Redmi Note 11 (Snapdragon 680)','N/A','1. Kill app\n2. Launch\n3. Measure','Time to first frame < 4 s','','Not Run','','','Yes'),
    ('PDF Export','PDF Generation Time – Simple','Simple resume PDF generated in <3 s','High','Resume with 3 sections','N/A','1. Tap Preview\n2. Measure render time','PDF preview appears in < 3 s','','Not Run','','','Yes'),
    ('PDF Export','PDF Generation Time – Complex','Full 10-section resume exported in <8 s','High','Resume with all 10 sections + photo','N/A','1. Tap Download PDF\n2. Measure','PDF file saved in < 8 s','','Not Run','','','Yes'),
    ('AI Response','AI Suggestion Response Time','AI returns result in <10 s on good network','High','4G/WiFi; AI configured','Prompt: moderate length','1. Trigger AI enhance\n2. Measure API call time','Response < 10 s on good network','','Not Run','','','Yes'),
    ('AI Response','AI Response Time – Slow Network','AI times out gracefully on slow network','Medium','Network throttled to 3G','Prompt: short','1. Throttle network\n2. Trigger AI call','Timeout message after 30 s; no hang','','Not Run','','','Yes'),
    ('Memory','Memory Under 150 MB Normal Use','App heap stays under 150 MB during resume editing','High','Android Studio Profiler','Normal use for 10 min','1. Profile heap during editing session','Max heap < 150 MB','','Not Run','','','Yes'),
    ('Memory','No Leak After Template Cycle','Switching 10 templates does not grow heap','High','Android Studio Profiler','N/A','1. Cycle through 10 templates\n2. Monitor heap','No unbounded heap growth','','Not Run','','','Yes'),
    ('Render','Scroll Performance – 50 Resumes','50-resume list scrolls at 60 fps','Medium','50 resumes in list','N/A','1. Scroll fast through 50 resumes\n2. Check with GPU profiler','No jank; 60 fps maintained','','Not Run','','','Yes'),
    ('Render','Preview Scroll Performance','PDF preview scrolls smoothly','Medium','Complex PDF preview','N/A','1. Open multi-page preview\n2. Scroll fast','Smooth at 60 fps; no dropped frames','','Not Run','','','Yes'),
    ('Network','Backup Upload Time – 10 Resumes','Backup 10 resumes in <15 s on WiFi','High','WiFi; Supabase configured','10 resumes each ~50 KB','1. Backup with 10 resumes\n2. Measure time','Upload complete in < 15 s','','Not Run','','','Yes'),
    ('Network','Restore Download Time – 10 Resumes','Restore 10 resumes in <15 s on WiFi','High','WiFi; backup in Supabase','10 resumes','1. Restore\n2. Measure','Download + parse complete in < 15 s','','Not Run','','','Yes'),
    ('Battery','Session Battery Use','1-hour active session uses <5% battery','Medium','Fully charged device; battery stats','N/A','1. Use app actively 1 hr\n2. Check battery drain','<5% drain per hour active use','','Not Run','','','No'),
    ('Concurrency','Simultaneous PDF + AI','PDF export and AI call at same time','Medium','Normal device','N/A','1. Start PDF export\n2. Immediately trigger AI call','Both complete correctly; no crash or starvation','','Not Run','','','Yes'),
    ('DB Performance','Hive Read – 100 Resumes','Hive read of 100 resumes in <500 ms','High','100 resumes in Hive','N/A','1. Pre-populate 100 resumes\n2. Load list\n3. Measure','List loads in < 500 ms','','Not Run','','','Yes'),
]

# ── COMPATIBILITY ──────────────────────────────────────────────────────────────
COMPATIBILITY = [
    ('OS','Android 10 (API 29)','All features work on Android 10','High','Android 10 emulator','N/A','1. Run full smoke suite on Android 10','All smoke tests pass','','Not Run','','','No'),
    ('OS','Android 11 (API 30)','All features work on Android 11','High','Android 11 device or emulator','N/A','1. Run smoke suite','All smoke tests pass','','Not Run','','','No'),
    ('OS','Android 12 (API 31-32)','All features work on Android 12','High','Android 12 emulator','N/A','1. Run smoke suite','All smoke tests pass','','Not Run','','','No'),
    ('OS','Android 13 (API 33)','All features work on Android 13','High','Android 13 device','N/A','1. Run smoke suite','All smoke tests pass','','Not Run','','','No'),
    ('OS','Android 14 (API 34)','All features work on Android 14','High','Android 14 device','N/A','1. Run smoke suite','All smoke tests pass','','Not Run','','','No'),
    ('Screen','5-inch Phone (1080×1920)','UI renders on small screen','High','5-inch phone emulator','N/A','1. Run app\n2. Check all screens','No overflow; all content readable','','Not Run','','','No'),
    ('Screen','6.7-inch Phone (1440×3200)','UI on large phone correct','Medium','6.7-inch phone','N/A','1. Run app','Proper scaling; no excessive whitespace','','Not Run','','','No'),
    ('Screen','10-inch Tablet (1280×800)','Tablet layout adapts correctly','Medium','10-inch tablet emulator','N/A','1. Run app on tablet','Two-column or wider layout used; no single-column stretch','','Not Run','','','No'),
    ('DPI','ldpi (120 dpi)','App renders at low DPI','Low','ldpi emulator','N/A','1. Run on ldpi emulator','Content visible; no bitmap overflow','','Not Run','','','No'),
    ('DPI','xxxhdpi (640 dpi)','App renders at ultra-high DPI','Medium','xxxhdpi device (Pixel 7 Pro)','N/A','1. Run on xxxhdpi','Icons and text sharp; no blurriness','','Not Run','','','No'),
    ('Manufacturer','Samsung OneUI','No OneUI-specific rendering issues','High','Samsung Galaxy device','N/A','1. Run full smoke on Samsung device','All screens render correctly under OneUI','','Not Run','','','No'),
    ('Manufacturer','Xiaomi MIUI','No MIUI-specific background kill or UI issues','High','Xiaomi device; allow background activity','N/A','1. Run app\n2. Background for 5 min\n3. Reopen','App not killed; state preserved','','Not Run','','','No'),
    ('Manufacturer','OnePlus OxygenOS','No OxygenOS-specific issues','Medium','OnePlus device','N/A','1. Run core flows','Core flows work; no OxygenOS-specific crash','','Not Run','','','No'),
    ('Language','English Locale','All strings correct in English','High','Device language: English','N/A','1. Set locale English\n2. Run app','All text in English; no missing strings','','Not Run','','','No'),
    ('Language','Spanish Locale','No string overflow in Spanish','Medium','Device language: Spanish','N/A','1. Set locale Spanish\n2. Run app','UI adapts; no truncated labels','','Not Run','','','No'),
    ('Language','Hindi Locale','Devanagari script renders correctly','Medium','Device language: Hindi','N/A','1. Set locale Hindi\n2. Run app','Hindi system strings render; app strings in English (no i18n yet)','','Not Run','','','No'),
    ('Pixel Density','Font Scale – Largest','UI adapts when system font size = largest','Medium','System font size: Largest','N/A','1. Set font size to largest\n2. Navigate app','Labels not truncated; layout adapts','','Not Run','','','No'),
    ('Night Mode','Force Dark Mode (system)','System force-dark does not break custom theme','Medium','Developer options: Force dark mode','N/A','1. Enable force dark\n2. Run app','App uses own dark theme; no double-dark artefacts','','Not Run','','','No'),
]

# ── ACCESSIBILITY ─────────────────────────────────────────────────────────────
ACCESSIBILITY = [
    ('TalkBack','Home Screen TalkBack Navigation','All cards announced by TalkBack','High','TalkBack enabled','N/A','1. Enable TalkBack\n2. Swipe through Home tab','Each resume card announced with title and date','','Not Run','','','No'),
    ('TalkBack','Resume Editor TalkBack','All section labels read aloud','High','TalkBack enabled','N/A','1. Open resume editor\n2. Swipe through sections','Section names announced correctly','','Not Run','','','No'),
    ('TalkBack','Buttons Have Semantic Labels','All icon buttons have accessibility labels','High','TalkBack; code review','N/A','1. TalkBack on\n2. Focus each icon button','Label announced e.g. "Add experience", not "Button"','','Not Run','','','No'),
    ('TalkBack','Dialog Buttons Announced','Confirm/Cancel in dialogs readable','High','TalkBack enabled','N/A','1. Trigger delete confirm dialog\n2. Listen for announcements','Both "Cancel" and "Delete All" announced correctly','','Not Run','','','No'),
    ('Colour Contrast','WCAG AA Normal Text','Text ≥4.5:1 contrast ratio','High','Colour contrast analyser tool','N/A','1. Screenshot each main screen\n2. Run contrast check on body text','All normal text passes 4.5:1 WCAG AA ratio','','Not Run','','','No'),
    ('Colour Contrast','WCAG AA Large Text','Large text ≥3:1 contrast ratio','Medium','Colour contrast tool','N/A','1. Check section headers','Headers pass 3:1 contrast ratio','','Not Run','','','No'),
    ('Colour Contrast','Dark Mode Contrast Maintained','Contrast ratios maintained in dark mode','High','Dark mode enabled; contrast tool','N/A','1. Enable dark mode\n2. Run contrast checks','All body text still passes WCAG AA in dark mode','','Not Run','','','No'),
    ('Touch Targets','All Tappable Elements ≥48dp','No small tap targets','High','Layout inspector','N/A','1. Inspect all interactive elements with layout inspector','All interactive elements ≥48×48 dp','','Not Run','','','No'),
    ('Focus Order','Keyboard/D-Pad Focus Order','Focus moves logically top-to-bottom','Medium','Bluetooth keyboard or D-pad emulator','N/A','1. Use Tab/D-pad to navigate','Focus moves in reading order; no skipped elements','','Not Run','','','No'),
    ('Text Scaling','125% Font Scale','UI readable at 125% system font','High','System font scale: 125%','N/A','1. Set font scale 125%\n2. Navigate app','No text overlap; labels not clipped','','Not Run','','','No'),
    ('Text Scaling','150% Font Scale','UI readable at 150% system font','High','System font scale: 150%','N/A','1. Set font scale 150%\n2. Navigate app','Scrollable if needed; no content lost','','Not Run','','','No'),
    ('Error Messages','Error Messages Are Descriptive','Errors describe problem and action','High','Various error conditions','N/A','1. Trigger network error, OTP error, payment error','Each error message tells user what happened and next step','','Not Run','','','No'),
    ('Motion','Reduce Motion Respected','Animations suppressed when reduce motion on','Medium','Device reduce motion: ON','N/A','1. Enable Reduce Motion in accessibility\n2. Open app','Animations minimal or absent; no looping effects','','Not Run','','','No'),
    ('Images','Profile Photo Alt Text','Profile image has accessibility description','Low','TalkBack enabled','Photo set','1. Set profile photo\n2. TalkBack focus on photo','Announced as "Profile photo" not just "Image"','','Not Run','','','No'),
    ('Headings','Semantic Headings Used','Section titles marked as headings for TalkBack','Medium','TalkBack heading navigation','N/A','1. TalkBack headings shortcut\n2. Navigate','Jumps to section headings correctly','','Not Run','','','No'),
]

# ── NEGATIVE ──────────────────────────────────────────────────────────────────
NEGATIVE = [
    ('Authentication','Login with Empty Phone Number','Submit with empty phone field','High','Login screen','Empty','1. Tap phone field\n2. Leave empty\n3. Tap Send OTP','Validation error: "Phone number required"','','Not Run','','','Yes'),
    ('Authentication','Login with Invalid Format Phone','Enter non-numeric phone','High','Login screen','Value: "abcdefg"','1. Type letters in phone field\n2. Tap Send OTP','Validation error: "Enter a valid phone number"','','Not Run','','','Yes'),
    ('Authentication','Wrong OTP 5 Times','Enter wrong OTP repeatedly','High','OTP verify screen','OTP: 000000 × 5','1. Enter wrong OTP 5 times','Account temporarily locked; error message shown','','Not Run','','','Yes'),
    ('Resume Editor','Save with Empty Required Fields','Save personal info with no name','High','Personal info screen','Name: ""','1. Clear name field\n2. Tap Save','Error: "Name is required"; save blocked','','Not Run','','','Yes'),
    ('Resume Editor','Email Field – Invalid Format','Save with malformed email','High','Personal info screen','Email: "notanemail"','1. Enter invalid email\n2. Save','Validation error: "Enter a valid email address"','','Not Run','','','Yes'),
    ('Resume Editor','Phone Field – Symbols Only','Enter symbols in phone field','Medium','Personal info screen','Value: "!@#$%^&*()"','1. Enter symbols in phone field\n2. Save','Validation error or symbols stripped','','Not Run','','','Yes'),
    ('PDF Export','Export with Empty Resume','Export PDF with no content','High','New blank resume','N/A','1. Create blank resume\n2. Preview\n3. Download PDF','Warning shown: "Add some content before exporting"','','Not Run','','','Yes'),
    ('AI','AI with Extremely Long Input','Send 10,000-char prompt to AI','Medium','AI configured','Input: 10,000 characters','1. Paste 10,000 chars in enhancer\n2. Enhance','Input truncated or error message; no crash','','Not Run','','','Yes'),
    ('AI','AI Call When Offline','Trigger AI when no network','High','Airplane mode','N/A','1. Enable airplane mode\n2. Use AI tool','Friendly error: "No internet connection"; no crash','','Not Run','','','Yes'),
    ('Backup','Backup with No Resumes','Backup when local storage is empty','Medium','No resumes created','N/A','1. Delete all resumes\n2. Tap Backup','Message: "Nothing to back up"; no crash','','Not Run','','','Yes'),
    ('Backup','Restore with Wrong Sync Code','Restore with a code that has no data','High','Backup sheet','Sync code: "code-that-doesnt-exist"','1. Enter non-existent sync code\n2. Restore','Message: "No backup found for this code"','','Not Run','','','Yes'),
    ('Subscription','Purchase with Insufficient Balance','Test card for insufficient funds','High','Razorpay test mode','Test card: insufficient funds scenario','1. Select plan\n2. Use failure test card','Payment failure message shown; plan not upgraded','','Not Run','','','Yes'),
    ('Template','Apply Template with No Resume Content','Template applied to empty resume','Medium','Empty resume','N/A','1. Select any template on blank resume','Template applied; empty preview shown without crash','','Not Run','','','Yes'),
    ('LinkedIn Import','Import Empty Text','Submit empty LinkedIn import','High','LinkedIn import screen','Input: ""','1. Leave input empty\n2. Tap Import','Error: "Please paste your LinkedIn profile"','','Not Run','','','Yes'),
    ('Job Tracker','Add Job with Empty Company','Save job entry with empty company name','High','Job tracker','Company: ""','1. Tap Add Job\n2. Leave company empty\n3. Save','Validation error: "Company is required"','','Not Run','','','Yes'),
    ('Settings','Delete All with No Data','Delete All when nothing saved','Low','No resumes saved','N/A','1. Settings → Delete All\n2. Confirm','Success message or "Nothing to delete"; no crash','','Not Run','','','Yes'),
    ('Profile Photo','Upload Corrupt Image','Upload a corrupt/invalid image file','Medium','Photo upload screen','File: corrupt.jpg (0 bytes)','1. Select corrupt file as profile photo','Error: "Could not load image"; no crash; photo not updated','','Not Run','','','Yes'),
    ('Resume','Duplicate with Deleted Original','Duplicate a resume after original deleted','Low','Resume list','N/A (edge case)','1. Delete resume A while copy dialog open\n2. Confirm duplicate','Graceful error; no orphan data created','','Not Run','','','Yes'),
]

# ── BOUNDARY ──────────────────────────────────────────────────────────────────
BOUNDARY = [
    ('Resume Editor','Name – Exactly 1 Character','Min valid name is 1 character','High','Personal info screen','Name: "A"','1. Enter single character name\n2. Save','Saved successfully; renders in preview','','Not Run','','','Yes'),
    ('Resume Editor','Name – 255 Characters (Max)','Max valid name length','High','Personal info screen','Name: "A" × 255','1. Enter 255-char name\n2. Save','Saved; no truncation; renders in PDF','','Not Run','','','Yes'),
    ('Resume Editor','Name – 256 Characters (Over Max)','Over max name length rejected','High','Personal info screen','Name: "A" × 256','1. Enter 256-char name\n2. Observe','Input capped at 255 or validation error shown','','Not Run','','','Yes'),
    ('Resume Editor','Phone – 7 Digits (Min)','Minimum phone length accepted','Medium','Personal info screen','Phone: "1234567"','1. Enter 7-digit phone\n2. Save','Saved without error','','Not Run','','','Yes'),
    ('Resume Editor','Phone – 15 Digits (E.164 Max)','Max phone length accepted','Medium','Personal info screen','Phone: "+999999999999999"','1. Enter 15-digit phone\n2. Save','Saved without error','','Not Run','','','Yes'),
    ('Resume Editor','Phone – 16 Digits (Over Max)','Over-length phone rejected','Medium','Personal info screen','Phone: 16 digits','1. Enter 16-digit phone\n2. Save','Validation error or input capped','','Not Run','','','Yes'),
    ('Resume Editor','0 Experience Entries','Resume with no experience entries','Medium','Resume editor','N/A','1. Leave experience section empty\n2. Preview PDF','Experience section hidden or "Not provided"; no crash','','Not Run','','','Yes'),
    ('Resume Editor','10 Experience Entries','Max realistic experience entries','High','Resume editor','10 entries','1. Add exactly 10 experience entries\n2. Preview PDF','All 10 entries in PDF; layout intact','','Not Run','','','Yes'),
    ('Resume Editor','20 Skills','Large skills list handled','Medium','Skills screen','20 skill entries','1. Add 20 skills\n2. Preview','All 20 skills listed in preview; no overflow','','Not Run','','','Yes'),
    ('Resume Editor','Summary – Exactly 50 Characters','Short summary boundary','Low','Summary screen','50-char summary','1. Enter exactly 50 characters\n2. Save','Saved; visible in preview','','Not Run','','','Yes'),
    ('Resume Editor','Summary – 2000 Characters','Maximum practical summary length','Medium','Summary screen','2000-char summary','1. Enter 2000 characters\n2. Save\n3. Preview PDF','Text wraps correctly in PDF; no data loss','','Not Run','','','Yes'),
    ('AI','AI Prompt – 1 Character','Minimal prompt processed or error','Medium','AI enhancer','Prompt: "X"','1. Enter single character\n2. Enhance','Either meaningful error or graceful AI response','','Not Run','','','Yes'),
    ('AI','AI Prompt – 4096 Characters','Max token-length prompt','High','AI enhancer','4096-char prompt','1. Enter max-length prompt\n2. Enhance','Request sent and response returned; no crash','','Not Run','','','Yes'),
    ('Job Tracker','0 Jobs in Tracker','Empty tracker state handled','Low','Job tracker','No jobs added','1. Open empty job tracker','Empty state illustration shown; "Track your first job" CTA','','Not Run','','','Yes'),
    ('Job Tracker','100 Jobs in Tracker','Large job list performance','Medium','Job tracker','100 job entries','1. Add 100 jobs\n2. Scroll list','List scrolls smoothly; no ANR','','Not Run','','','Yes'),
    ('Subscription','Free Tier – AI Limit Boundary','AI usable up to free limit, blocked at limit+1','High','Free plan','4 prior AI uses (assuming limit=5)','1. Use AI on 5th use\n2. Attempt 6th','5th succeeds; 6th shows upgrade prompt','','Not Run','','','Yes'),
    ('Sync','Sync Code – Exactly 1 Character','Minimum sync code accepted','Medium','Backup sheet','Sync code: "a"','1. Set sync code "a"\n2. Backup','Backup succeeds with 1-char code','','Not Run','','','Yes'),
    ('Sync','Sync Code – 100 Characters','Very long sync code accepted','Low','Backup sheet','Sync code: 100-char string','1. Set 100-char sync code\n2. Backup','Backup succeeds; code not truncated','','Not Run','','','Yes'),
    ('Resume','0 Resumes in List','Empty home state','Low','Home tab','No resumes','1. Delete all resumes\n2. Open Home','Empty state shown with CTA','','Not Run','','','Yes'),
    ('Resume','50 Resumes in List','Performance at list boundary','Medium','Home tab','50 resumes','1. Create/import 50 resumes\n2. Open Home tab','List loads in <3 s; no crash','','Not Run','','','Yes'),
]

# ── SUBSCRIPTION / BILLING ────────────────────────────────────────────────────
SUBSCRIPTION = [
    ('Plans','View Subscription Plans','All plans displayed correctly','High','Upgrade screen','N/A','1. Tap Upgrade or locked feature','Monthly, Quarterly, Annual plans shown with prices','','Not Run','','','Yes'),
    ('Plans','Monthly Plan Price Correct','Monthly price matches Play Store listing','Critical','Upgrade screen; Play Store configured','N/A','1. Open upgrade screen\n2. Note monthly price','Price matches Play Console product setup','','Not Run','','','Yes'),
    ('Plans','Annual Plan Discount Shown','Annual plan shows savings %','High','Upgrade screen','N/A','1. Open upgrade screen\n2. Check annual plan','"Save X%" badge visible on annual plan','','Not Run','','','Yes'),
    ('Payment','Monthly Plan Purchase','Complete monthly subscription purchase','Critical','Google Play test account; test product','Test product: monthly_premium','1. Select Monthly\n2. Tap Subscribe\n3. Confirm in Play','Subscription active; premium features unlocked','','Not Run','','','Yes'),
    ('Payment','Annual Plan Purchase','Complete annual subscription purchase','Critical','Google Play test account','Test product: annual_premium','1. Select Annual\n2. Tap Subscribe\n3. Confirm','Subscription active; premium badge shown','','Not Run','','','Yes'),
    ('Payment','Razorpay Payment Success','Razorpay checkout completes and unlocks premium','Critical','Razorpay test key; test card','Card: 4111 1111 1111 1111','1. Tap Subscribe (Razorpay flow)\n2. Enter test card\n3. Complete','Premium activated; plan badge in settings','','Not Run','','','Yes'),
    ('Payment','Razorpay Payment Failure','Payment failure shows error without upgrade','High','Razorpay test key; failure card','Failure test card','1. Select plan\n2. Enter failure card\n3. Submit','Error message; plan remains free','','Not Run','','','Yes'),
    ('Payment','Payment Dismissed','Closing Razorpay sheet keeps free plan','High','Razorpay test mode','N/A','1. Open Razorpay\n2. Close sheet','Plan remains free; no phantom subscription','','Not Run','','','Yes'),
    ('Feature Gates','AI Cover Letter Locked on Free','Cover letter inaccessible on free plan','Critical','Free plan account','N/A','1. Tap AI Cover Letter','Upgrade prompt shown; cover letter blocked','','Not Run','','','Yes'),
    ('Feature Gates','Skill Analyzer Locked on Free','Skill analyzer inaccessible on free plan','High','Free plan account','N/A','1. Tap Skill Analyzer','Upgrade prompt shown','','Not Run','','','Yes'),
    ('Feature Gates','Cloud Sync Locked on Free','Backup & Sync locked on free plan','High','Free plan account','N/A','1. Tap Backup & Sync','Premium required message shown','','Not Run','','','Yes'),
    ('Feature Gates','Priority Support Locked on Free','Priority support locked on free plan','Medium','Free plan account','N/A','1. Tap Priority Support','Upgrade prompt shown','','Not Run','','','Yes'),
    ('Feature Gates','AI Tools Unlocked on Premium','All AI tools accessible on premium','Critical','Premium plan active','N/A','1. Open AI Assistant\n2. Check all tool cards','No lock icons; all tools accessible','','Not Run','','','Yes'),
    ('Restoration','Plan Restored After Reinstall','Previous subscription restored on new install','Critical','Premium account; app reinstalled','Google account','1. Reinstall app\n2. Login with same Google account','Premium restored via Play Billing; features unlocked','','Not Run','','','Yes'),
    ('Restoration','Plan Restored on New Device','Premium active on second device','High','Premium account; new device','Same Google account','1. Log in on new device','Premium subscription restored','','Not Run','','','Yes'),
    ('Expiry','Expired Subscription Reverts to Free','Features locked after plan expires','Critical','Expired test subscription','N/A','1. Let subscription expire\n2. Open locked feature','Feature locked; upgrade prompt shown','','Not Run','','','Yes'),
    ('Receipt','Payment Receipt Accessible','Users can view transaction receipt','Medium','Completed payment','N/A','1. Complete payment\n2. Check Settings or email','Receipt or confirmation accessible via Play account','','Not Run','','','No'),
    ('Free Trial','Free Trial Unlocks Premium','Trial period grants full premium access','High','Free trial offer visible','N/A','1. Accept free trial\n2. Access cover letter','Premium features accessible during trial','','Not Run','','','Yes'),
    ('Free Trial','Free Trial Converts to Paid','Trial auto-converts unless cancelled','High','Trial started >7 days ago','N/A','1. Let trial expire without cancelling','Billing begins; premium remains active','','Not Run','','','No'),
    ('Pricing Region','Correct Currency for Region','Pricing shown in local currency','High','Device locale: India / USA / UK','N/A','1. Set locale\n2. Open upgrade screen','INR / USD / GBP shown appropriately','','Not Run','','','Yes'),
    ('Premium Badge','Premium Badge Displayed','Settings shows premium status','High','Premium plan active','N/A','1. Open Settings\n2. Check subscription section','"Premium" badge or "Active" status shown','','Not Run','','','Yes'),
]

# ── BACKUP & SYNC ─────────────────────────────────────────────────────────────
BACKUP_SYNC = [
    ('Setup','Set Sync Code','User can set a sync code','High','Settings → Backup & Sync','Code: "my-sync-2025"','1. Open Backup & Sync\n2. Tap "Set Code"\n3. Enter code\n4. Save','Code saved; "Code: my-sync-2025" shown in green chip','','Not Run','','','Yes'),
    ('Setup','Change Sync Code','User can update existing code','High','Sync code already set','New code: "new-sync-001"','1. Tap "Change"\n2. Enter new code\n3. Save','New code active; old code replaced','','Not Run','','','Yes'),
    ('Setup','Clear Sync Code','User can clear sync code to device-only','Medium','Sync code set','N/A','1. Tap "Clear" next to code','Code removed; device-only warning shown','','Not Run','','','Yes'),
    ('Backup','Backup Resumes to Cloud','Resumes uploaded to Supabase','Critical','Resumes exist; internet; sync code set','3 resumes','1. Tap "Backup to Cloud"','Success message: "3 resume(s) saved to code …"','','Not Run','','','Yes'),
    ('Backup','Backup Job Tracker to Cloud','Job tracker data uploaded','High','Jobs in tracker; sync code set','5 jobs','1. Tap "Backup to Cloud"','Success message includes job tracker count','','Not Run','','','Yes'),
    ('Backup','Backup Empty Data','Nothing to back up message','Medium','No resumes or jobs','N/A','1. Tap Backup with empty data','Message: "Nothing to back up yet"','','Not Run','','','Yes'),
    ('Backup','Backup While Offline','Network unavailable during backup','High','Airplane mode on','N/A','1. Enable airplane mode\n2. Tap Backup','Error: "Check your internet connection"','','Not Run','','','Yes'),
    ('Backup','Backup Progress Shown','Spinner during backup','Medium','Normal network','N/A','1. Start backup\n2. Watch UI during operation','Loading spinner visible; buttons disabled while in progress','','Not Run','','','Yes'),
    ('Restore','Restore from Same Device Code','Data restored from own backup','Critical','Backup exists; same sync code','Sync code: "my-sync-2025"','1. Tap Restore\n2. Confirm','All backed-up resumes restored; job tracker restored','','Not Run','','','Yes'),
    ('Restore','Restore from Different Device','Resume from Device A appears on Device B','Critical','Backup from Device A; same sync code on Device B','N/A','1. On Device B: set same code\n2. Restore','Device A\'s resumes appear on Device B','','Not Run','','','Yes'),
    ('Restore','Restore Merges with Existing','Newer local items not overwritten','High','Local item newer than cloud item','Local resume last edited 1 hr ago; cloud 2 days ago','1. Restore when local is newer','Local newer resume kept; older cloud version merged without overwriting','','Not Run','','','Yes'),
    ('Restore','Restore No Backup Found','Correct message for missing backup','High','Wrong sync code','Code: "wrong-code-xyz"','1. Enter wrong code\n2. Restore','Message: "No cloud backup was found for this code"','','Not Run','','','Yes'),
    ('Restore','Restore Offline','No network during restore','High','Airplane mode','N/A','1. Airplane mode\n2. Restore','Error: "Check your internet connection"','','Not Run','','','Yes'),
    ('Restore','Restore Updates UI Immediately','My Resumes tab refreshes after restore','High','Resumes restored','N/A','1. Restore\n2. Navigate to Home tab','New resumes visible immediately without manual refresh','','Not Run','','','Yes'),
    ('Integrity','Restored Resume Opens Correctly','Restored resume fully editable','Critical','Restored resume','N/A','1. Restore\n2. Open a restored resume\n3. Edit','All fields present and editable; PDF exports correctly','','Not Run','','','Yes'),
    ('Integrity','Job Tracker Data Accurate After Restore','Job status and dates correct post-restore','High','Job tracker backed up and restored','N/A','1. Backup with 5 jobs\n2. Delete local\n3. Restore','All 5 jobs with correct statuses restored','','Not Run','','','Yes'),
    ('Concurrency','Simultaneous Backup from Two Devices','Two devices backup simultaneously','Medium','Two devices; same sync code','N/A','1. Backup from Device A\n2. Immediately backup from Device B','No data corruption; both backups recorded or last-write-wins gracefully','','Not Run','','','No'),
    ('Data Safety','Backup Does Not Duplicate Resumes','Repeated backup does not create duplicates','High','Backup run twice without changes','N/A','1. Backup\n2. Restore\n3. Backup again\n4. Restore again','Resume count stays constant; no duplicates','','Not Run','','','Yes'),
    ('Premium Gate','Backup Locked for Free Users','Backup & Sync requires premium','High','Free plan account','N/A','1. Open Backup & Sync on free plan','Upgrade prompt shown; backup/restore buttons not active','','Not Run','','','Yes'),
    ('Sync Code Isolation','Different Codes = Different Data','Code A data not accessible via code B','Critical','Two sync codes with different backups','Code A data; Code B data','1. Backup on Code A\n2. Switch to Code B\n3. Restore','Only Code B data returned; Code A data not exposed','','Not Run','','','Yes'),
]

# ── CROSS-PLATFORM ────────────────────────────────────────────────────────────
CROSS_PLATFORM = [
    ('Android','Resume Created on Android Opens on Web','Resume data transferable to web via sync','High','Supabase sync; web app available','Android + Chrome web build','1. Create resume on Android\n2. Backup\n3. Open web app\n4. Restore','Same resume visible on web app','','Not Run','','','No'),
    ('Android','PDF Export on Android 10','PDF exports on oldest supported Android','High','Android 10 device','N/A','1. Export PDF on Android 10','PDF generated and saved','','Not Run','','','No'),
    ('Android','PDF Export on Android 14','PDF exports on latest Android','High','Android 14 device','N/A','1. Export PDF on Android 14','PDF generated; saved to correct location per Android 14 scoped storage','','Not Run','','','No'),
    ('Android – Web','Backup on Android → Restore on Chrome Web','Cross-platform sync works','High','Both Android and web app; Supabase','Sync code shared across platforms','1. Backup on Android with sync code\n2. Open web app\n3. Restore','Web app shows resumes from Android backup','','Not Run','','','No'),
    ('Chrome Web','Web App Loads Correctly','Web version runs in Chrome','High','Web build deployed','Chrome 120+','1. Open web URL in Chrome\n2. Navigate all sections','All main features accessible; no console errors','','Not Run','','','No'),
    ('Chrome Web','Web PDF Export','PDF downloadable from web version','High','Web build; Chrome','N/A','1. Open resume on web\n2. Preview\n3. Download PDF','PDF downloaded via browser download','','Not Run','','','No'),
    ('Chrome Web','Web Responsive Design','Web app responsive at 1280 × 800','High','Desktop browser; 1280px width','N/A','1. Open at 1280px\n2. Navigate','Layout uses desktop breakpoints; no horizontal scroll','','Not Run','','','No'),
    ('Chrome Web','Web at Mobile Width (375px)','Web app usable at mobile viewport','Medium','Browser dev tools; 375px width','N/A','1. Resize to 375px\n2. Navigate','Responsive layout; bottom nav or hamburger','','Not Run','','','No'),
    ('Android vs Web','Template Renders Same Cross-Platform','Same template looks identical on Android and web','High','Android + web build; same resume','Template: "Modern Edge"','1. Select Modern Edge on Android\n2. Preview PDF\n3. Open same resume on web\n4. Preview PDF','PDF layout visually consistent across platforms','','Not Run','','','No'),
    ('Android – Small Device','Data from Large Screen Renders on Small Phone','Data from tablet or web renders on 5-inch phone','Medium','5-inch phone; resume with lots of content','N/A','1. Create complex resume on 6.7-inch phone\n2. Open on 5-inch phone','Editor and preview accessible without data loss','','Not Run','','','No'),
    ('Locale Cross-Platform','Locale Settings Consistent','Same app behaviour across Android locales','Medium','English and Spanish locales','N/A','1. Test on English Android\n2. Test on Spanish Android','No feature discrepancy; same core flows work','','Not Run','','','No'),
    ('Night Mode','Dark Mode on Android 12 vs Android 10','Dark mode consistent across OS versions','Medium','Android 10 + Android 12','N/A','1. Enable dark mode on both\n2. Compare UI','Consistent dark theme; no white flash or broken colors','','Not Run','','','No'),
    ('Google Account','Same Account on Two Android Devices','Account data consistent on two devices','High','Same Google account; two Android devices','N/A','1. Login on Device A\n2. Login on Device B with same account\n3. Backup on A\n4. Restore on B','Resumes match; no data conflict','','Not Run','','','Yes'),
    ('Web Firebase Auth','Google Sign-In Works on Web Build','Firebase Auth works in web context','High','Web build; Firebase configured for web','Google account','1. Open web app\n2. Tap "Continue with Google"','Google OAuth completes; user logged in on web','','Not Run','','','No'),
    ('Orientation','Portrait to Landscape on Android','Rotation handled on Android','Medium','Android device','N/A','1. Open editor\n2. Rotate to landscape','Layout adapts; content not lost','','Not Run','','','No'),
    ('File Picker','File Picker Works on Different Android Versions','File picker compatible with Android 10-14','High','Android 10, 12, 14 devices','N/A','1. Upload profile photo on each version','Photo picker opens; photo selected and displayed','','Not Run','','','Yes'),
    ('Background Sync','App Backgrounded During Sync on Different OS Versions','Sync completes even with aggressive background kill','Medium','Android 12+ (background restrictions)','N/A','1. Start backup\n2. Immediately background app','Backup completes or resumes; not silently killed','','Not Run','','','No'),
    ('Notification','Push Notification Delivery Cross-Version','Notifications work on Android 10-14','Low','FCM configured; test notification sent','N/A','1. Send test push\n2. Check on Android 10, 12, 14','Notification received and tappable on all versions','','Not Run','','','No'),
]

# ── EXTRA FUNCTIONAL (extending coverage to ATS, Versioning, Portfolio, Tools)
FUNCTIONAL_EXTRA = [
    ('ATS Score','ATS Score Displayed','Resume ATS score shown after analysis','High','Resume with content; ATS feature enabled','N/A','1. Open resume\n2. Tap ATS Analyse\n3. Wait','Score 0-100 displayed; colour-coded indicator','','Not Run','','','Yes'),
    ('ATS Score','ATS Missing Keywords','Missing keywords highlighted','High','ATS feature enabled','Job description pasted','1. Run ATS analysis with JD\n2. Check keywords section','Keywords not in resume highlighted in red/orange','','Not Run','','','Yes'),
    ('ATS Score','ATS Score Improvements Applied','Score improves after suggestion applied','High','ATS score displayed; at least one suggestion','N/A','1. Apply top ATS suggestion\n2. Re-run analysis','Score higher than before by at least 5 points','','Not Run','','','Yes'),
    ('Resume Versioning','Version History Listed','Previous resume versions accessible','Medium','Resume edited 3+ times','N/A','1. Open resume version history','3 prior versions listed with timestamps','','Not Run','','','Yes'),
    ('Resume Versioning','Restore Previous Version','Older version can be restored','Medium','Version history with 2+ versions','N/A','1. Open history\n2. Select version 1\n3. Restore','Resume reverted to version 1 content','','Not Run','','','Yes'),
    ('Resume Versioning','Version Diff Preview','User can compare two versions','Low','Version history with 2 versions','N/A','1. Open history\n2. Tap "Compare"','Diff view shows added/removed sections','','Not Run','','','No'),
    ('Portfolio','Portfolio Generated from Resume','Portfolio page built from resume data','Medium','Premium; resume with content','N/A','1. Tap "Generate Portfolio"\n2. Confirm','Portfolio page created with sections from resume','','Not Run','','','No'),
    ('Portfolio','Portfolio URL Shared','Portfolio link shareable','Medium','Portfolio generated','N/A','1. Generate portfolio\n2. Tap "Share Link"','Share sheet with URL opens','','Not Run','','','No'),
    ('AI – RAOE2','Auto-Optimize Resume','RAOE2 rewrites resume for job description','High','Premium; resume and JD text','JD: Senior Product Manager','1. Open RAOE2\n2. Paste resume and JD\n3. Optimize','Rewritten resume with keyword insertions shown','','Not Run','','','Yes'),
    ('AI – RAOE2','RAOE2 Before/After Preview','Side-by-side comparison shown','High','RAOE2 optimization done','N/A','1. Run RAOE2\n2. View comparison','Before and after versions shown side-by-side','','Not Run','','','No'),
    ('AI – Roast Resume','Roast Resume Feature','AI provides humorous critique of resume','Low','AI configured; resume with content','N/A','1. Tap "Roast My Resume"','Entertaining critique generated; no sensitive data leaked','','Not Run','','','No'),
    ('AI – Skill Suggestions','Skill Suggestions by Role','Role-based skills suggested','Medium','AI configured','Role: "Full Stack Developer"','1. Tap Skill Suggestions\n2. Enter role','10+ relevant skills listed for role','','Not Run','','','Yes'),
    ('Career Tools','Career Path Screen','Career progression paths shown','Low','Career Tools tab','N/A','1. Tap "Career Path"','Career level progression displayed for a sample role','','Not Run','','','No'),
    ('Career Tools','Job Search Screen','Job search results displayed','Medium','Internet available','Search: "Flutter developer remote"','1. Tap "Job Search"\n2. Enter query','Job listings or deep link to job site opens','','Not Run','','','No'),
    ('Career Tools','Career Articles Shown','Career advice articles loaded','Low','Internet available','N/A','1. Tap "Career Articles"','List of articles with thumbnails shown','','Not Run','','','No'),
    ('Career Tools','Interview Prep Questions','Role-specific questions generated','Medium','AI configured','Role: "Data Analyst"','1. Tap "Interview Prep"\n2. Enter role\n3. Generate','At least 10 relevant questions shown','','Not Run','','','Yes'),
    ('Resume Import','Import Resume from File','Resume imported from JSON/text file','Medium','File available on device','resume.json file','1. Tap "Import Resume"\n2. Select file','Resume sections populated from file','','Not Run','','','No'),
    ('Resume Quality','Quality Score Displayed','Resume quality score visible in editor','Medium','Resume with content','N/A','1. Open resume in editor\n2. Look for quality indicator','Quality score shown (e.g. 72/100)','','Not Run','','','Yes'),
    ('Resume Quality','Quality Score Improves on Content Add','Adding content raises quality score','Medium','Resume with low score','Score < 50%','1. Note score\n2. Add 3 experience bullets\n3. Re-check score','Score increases after content added','','Not Run','','','Yes'),
    ('Notification','Rate App Prompt After 5 Uses','Rate app prompt shown after 5 app launches','Low','5 prior app launches','N/A','1. Open app on 5th launch','Rate app dialog or prompt shown','','Not Run','','','No'),
    ('Data Export','Export Resume as DOCX (if supported)','Resume exported as Word document','Medium','DOCX export feature enabled','N/A','1. Open resume\n2. Export → DOCX','Valid .docx file saved to device','','Not Run','','','No'),
    ('Translation','Resume Translated to Spanish','Resume content translated to Spanish','Medium','LibreTranslate or similar configured','Resume in English','1. Tap "Translate"\n2. Select Spanish','Resume sections translated; English original preserved','','Not Run','','','No'),
    ('Translation','Resume Translated to French','Resume content translated to French','Low','Translation service configured','Resume in English','1. Tap "Translate"\n2. Select French','French translation generated correctly','','Not Run','','','No'),
    ('Settings','Rate App via Settings','Tapping Rate App navigates to store','Low','Internet available','N/A','1. Settings → Rate App','Play Store / App Store opens on app page','','Not Run','','','No'),
    ('Settings','Share App via Settings','Share App opens share sheet with URL','Low','Internet available','N/A','1. Settings → Share App','Share sheet with Play Store URL opens','','Not Run','','','No'),
    ('Settings','Contact Support via Settings','Contact opens email client','Medium','Email client installed','N/A','1. Settings → Contact Support','Email compose window with support address opens','','Not Run','','','No'),
    ('Settings','Privacy Policy Readable','Privacy policy text fully accessible','Medium','Settings → Privacy Policy','N/A','1. Tap Privacy Policy\n2. Scroll entire document','All sections readable; no truncated content','','Not Run','','','No'),
    ('Settings','Terms of Service Readable','Terms document fully accessible','Low','Settings → Terms of Service','N/A','1. Tap Terms of Service\n2. Scroll','All sections readable','','Not Run','','','No'),
    ('Dashboard','Resume Card Shows Last Edited Date','Last-edited timestamp accurate','Medium','Resume edited recently','Resume edited 2 min ago','1. Open Home tab\n2. Check resume card date','Shows "Just now" or recent timestamp','','Not Run','','','Yes'),
    ('Dashboard','Resume Card Shows Template Name','Template name visible on card','Low','Resume with template selected','Template: "Corporate Navy"','1. Home tab\n2. Check resume card footer','Template name "Corporate Navy" shown on card','','Not Run','','','Yes'),
    ('Resume Editor','Reorder Experience Entries','Drag to reorder experience entries','Medium','2+ experience entries','N/A','1. Long-press experience entry\n2. Drag to new position','Entries reordered; new order reflected in preview','','Not Run','','','Yes'),
    ('Resume Editor','Reorder Education Entries','Drag to reorder education entries','Medium','2+ education entries','N/A','1. Long-press education entry\n2. Drag','Entries reordered correctly','','Not Run','','','Yes'),
    ('Resume Editor','Add Link to Experience','URL added to experience entry','Low','Experience entry exists','URL: "https://project.com"','1. Open experience entry\n2. Add URL field\n3. Save','URL shown in experience section of resume','','Not Run','','','Yes'),
    ('Preview','Two-Page Resume Preview','Long resume spans two pages in preview','High','Resume with 15+ entries','N/A','1. Expand resume with lots of content\n2. Preview','Preview shows page 1 and page 2 correctly','','Not Run','','','Yes'),
    ('Preview','Preview Page Navigation','User can navigate multi-page preview','Medium','Multi-page resume','N/A','1. Preview multi-page resume\n2. Scroll / swipe','All pages accessible in preview','','Not Run','','','Yes'),
    ('Preview','Zoom in Preview','User can pinch-to-zoom preview','Low','Preview screen open','N/A','1. Pinch to zoom in on preview','Preview zooms in without crash','','Not Run','','','No'),
    ('Onboarding','Onboarding Skip Works','Skip button bypasses all onboarding steps','High','Fresh install','N/A','1. Open fresh app\n2. Tap "Skip"','Redirected to login or home directly','','Not Run','','','Yes'),
    ('Onboarding','Onboarding Next/Back Navigation','User can navigate back through onboarding','Medium','Onboarding in progress','N/A','1. Advance to step 3\n2. Tap "Back"','Returns to step 2 with correct content','','Not Run','','','Yes'),
    ('Profile','Edit Account Display Name','User can update display name in profile','Medium','Profile screen accessible','New name: "Jane Doe"','1. Open Profile\n2. Edit display name\n3. Save','Display name updated in profile and settings','','Not Run','','','Yes'),
    ('Profile','Profile Avatar Shows Initials When No Photo','Fallback initials shown without photo','Low','No profile photo uploaded','Name: "John Doe"','1. No photo set\n2. Open profile avatar','Avatar shows "JD" initials','','Not Run','','','Yes'),
]

# ── EXTRA REGRESSION ───────────────────────────────────────────────────────────
REGRESSION_EXTRA = [
    ('Resume Editor','Multi-Line Bullet Points in PDF','Bullets with line breaks render in PDF','Medium','Experience entry with multi-line bullet','Text with \n','1. Add multi-line bullet\n2. Export PDF','Line breaks preserved in PDF','','Not Run','','','Yes'),
    ('Resume Editor','Emoji in Summary Field','Emojis stored and rendered correctly','Low','Summary screen','Summary with emoji 🚀','1. Enter summary with emoji\n2. Export PDF','Emoji appears in PDF as text fallback or renders','','Not Run','','','Yes'),
    ('Template','Template Switching Preserves Custom Colour','Custom colour retained on template change','Medium','Custom colour + template selected','Colour: #F39C12; Template: Modern Edge','1. Set colour\n2. Change template\n3. Preview','Custom colour applied to new template','','Not Run','','','Yes'),
    ('Resume Editor','Concurrent Edit and Preview','Edit while preview is open','Medium','Editor and preview open','N/A','1. Open resume\n2. Open preview in split mode\n3. Edit name','Preview updates with new name','','Not Run','','','Yes'),
    ('AI','AI Enhancer Retains Original on Cancel','Cancelling enhancement keeps original text','High','AI enhancer with text','Original: "managed projects"','1. Enter text\n2. Start enhancement\n3. Cancel','Original text unchanged after cancel','','Not Run','','','Yes'),
    ('Authentication','Login After Account Deletion','Deleted account prompts re-registration','High','Account deleted via data deletion','Deleted account credentials','1. Delete account\n2. Try to re-login\n3. Login with same phone','Fresh onboarding or account recreation flow triggered','','Not Run','','','Yes'),
    ('Career Tools','Job Tracker Notes Persist','Notes on job entries survive restart','High','Job entry with notes saved','Note: "Good company culture"','1. Add note to job\n2. Kill app\n3. Reopen','Note still present on job entry','','Not Run','','','Yes'),
    ('Settings','Theme Persists After Update','Theme setting unchanged after app update','High','Dark mode set; new build installed','N/A','1. Set dark mode\n2. Install new build','Theme still dark after update','','Not Run','','','Yes'),
    ('Resume Editor','Long URL in Link Field Renders in PDF','Very long hyperlink does not break layout','Low','Experience entry with URL','URL: 200-character URL','1. Add 200-char URL to experience\n2. Export PDF','URL truncated gracefully in PDF; layout intact','','Not Run','','','Yes'),
    ('Preview','Preview After Template Change Shows New Layout','Preview reflects template change immediately','High','Template changed just now','N/A','1. Change template\n2. Open preview','New template layout shown; old template gone','','Not Run','','','Yes'),
    ('Subscription','Premium Badge Gone After Downgrade','Premium badge hidden on free plan','High','Downgraded account','N/A','1. Downgrade plan\n2. Open Settings','No premium badge; free plan indicator shown','','Not Run','','','Yes'),
    ('Backup','Backup Code Case Insensitive Match','Sync code "TEST" same as "test"','Medium','Backup with code "TEST"','Restore code: "test"','1. Backup with "TEST"\n2. Restore with "test"','Data restored; codes treated as case-insensitive','','Not Run','','','Yes'),
    ('ATS','ATS Score Resets on Resume Change','ATS score cleared when resume edited','Medium','ATS score displayed','N/A','1. Run ATS\n2. Edit resume\n3. Check score area','Score shown as "Outdated – run again" or cleared','','Not Run','','','Yes'),
    ('Resume','Resume Renamed in Editor Reflects in List','Name change in editor shown on home card','High','Resume open in editor','New name: "Product Manager Resume"','1. Change resume title in editor\n2. Back to Home','Home card shows new name','','Not Run','','','Yes'),
    ('Deep Link','Deep Link to Specific Resume','Deep link opens correct resume directly','Medium','Resume ID known','Deep link: /resume/<id>','1. Trigger deep link with resume ID','Correct resume opens in editor','','Not Run','','','No'),
    ('AI','Bullet Generator Inserts into Correct Field','AI bullet inserted into selected experience entry only','High','2 experience entries; AI configured','N/A','1. Select Experience 2\n2. Generate bullet\n3. Insert','Bullet added to Experience 2 only; Experience 1 unchanged','','Not Run','','','Yes'),
    ('Performance','PDF Performance Not Degraded Post-Update','PDF generation time same as prior build','High','Performance baseline from prior build','Baseline: 3 s','1. Export PDF after new build\n2. Measure time','Time within 20% of baseline (≤3.6 s)','','Not Run','','','Yes'),
    ('Security','No Token in Crash Logs','Auth token absent from crash report','Critical','Crash reporting enabled (if any)','N/A','1. Trigger crash in test\n2. Check crash report','No plaintext token in crash metadata','','Not Run','','','No'),
    ('Translation','Translation Preserved on App Restart','Translated resume language retained after restart','Medium','Resume translated to Spanish','N/A','1. Translate resume\n2. Kill app\n3. Reopen','Spanish version still selected on reopen','','Not Run','','','Yes'),
    ('Navigation','Back from Deep Screen Returns to Correct Tab','Back nav returns to originating tab','High','Navigated from Career Tools to cover letter','N/A','1. From Career Tools, open cover letter\n2. Press Back','Returns to Career Tools tab, not Home','','Not Run','','','Yes'),
    ('Notification','No Duplicate Notifications','Backup success does not fire multiple toasts','Medium','Backup triggered','N/A','1. Tap Backup once\n2. Observe toasts','Exactly one success toast shown','','Not Run','','','Yes'),
    ('Resume Editor','Date Validation – From After To','Start date cannot be after end date','High','Experience date fields','From: 2024, To: 2022','1. Enter From > To dates\n2. Save','Validation error: "Start date must be before end date"','','Not Run','','','Yes'),
    ('Resume','Resume Count Updated on Delete','Home badge or count decrements on delete','Medium','3 resumes; resume count shown','N/A','1. Note count\n2. Delete one resume','Count decrements by 1 in Settings and Home','','Not Run','','','Yes'),
    ('Auth','Google Account Switch','Signing in with different Google account replaces session','High','Already logged in with Account A','Account B credentials','1. Logout\n2. Sign in with Account B','Account B session active; Account A data not accessible','','Not Run','','','Yes'),
    ('Data','Hive Box Not Corrupted After Force Close','App data intact after force kill','High','Resume in edit with unsaved changes','N/A','1. Edit resume\n2. Force kill app\n3. Reopen','Resume data intact up to last auto-save','','Not Run','','','Yes'),
    ('PDF','PDF Filename Contains Resume Name','Exported PDF filename uses resume title','Medium','Resume named "Senior Dev Resume"','N/A','1. Export PDF\n2. Check filename','Filename contains "Senior_Dev_Resume" or similar','','Not Run','','','Yes'),
    ('Settings','App Version Increments on New Build','About screen shows latest build number','High','New build installed','Build: 1780824930','1. Settings → About','Version shows latest build number','','Not Run','','','Yes'),
    ('Career Tools','Cover Letter Saves to Device','Generated cover letter can be saved','Medium','Cover letter generated','N/A','1. Generate cover letter\n2. Tap Save / Export','Text file or share sheet opens','','Not Run','','','No'),
    ('Home','Empty Search Shows No Results State','Search with no match shows empty state','Medium','Home tab; 3 resumes','Search: "ZZZZNOTEXISTS"','1. Search "ZZZZNOTEXISTS"','Empty state shown: "No resumes match your search"','','Not Run','','','Yes'),
    ('AI','AI Cover Letter Includes Job Title','Generated letter references job title','High','Premium; AI configured','Job: "Senior iOS Developer at Apple"','1. Generate cover letter for given job','Letter body mentions "Senior iOS Developer" and "Apple"','','Not Run','','','Yes'),
]

# ── EXTRA UI/UX ───────────────────────────────────────────────────────────────
UI_UX_EXTRA = [
    ('Layout','Card Shadow Consistent','All resume cards have same elevation/shadow','Low','Home tab; 3+ resumes','N/A','1. View resume cards','All cards same shadow depth; no inconsistency','','Not Run','','','No'),
    ('Layout','Section Dividers Visible in Editor','Section dividers clear in editor list','Medium','Editor open','N/A','1. Open resume editor\n2. Observe section separators','Thin dividers between sections visible','','Not Run','','','No'),
    ('Color','Primary Colour Consistent','Brand purple used consistently','Medium','Navigate all screens','N/A','1. Check all CTAs, FAB, active states','All primary actions use brand colour','','Not Run','','','No'),
    ('Icons','Iconsax Icons Consistent','All icons from same icon set','Low','All screens','N/A','1. Navigate all screens\n2. Check icon style','All icons consistent visual weight and style','','Not Run','','','No'),
    ('Feedback','Haptic Feedback on Important Actions','Vibration on delete/confirm','Low','Physical device; vibration on','N/A','1. Delete a resume\n2. Confirm','Short haptic vibration felt','','Not Run','','','No'),
    ('Typography','Heading Sizes Hierarchical','H1 > H2 > body text in size','Medium','All screens','N/A','1. Check screen titles, section headers, body','Clear size hierarchy maintained','','Not Run','','','No'),
    ('Typography','No Overflow in Narrow Containers','Text not clipped in narrow cards or chips','High','All screens with badges/chips','N/A','1. Check all badges and chips','No text clipped; ellipsis used if needed','','Not Run','','','No'),
    ('Loading','Skeleton Loader on Resume List','Skeleton shown while resumes load','Low','Slow device/network','N/A','1. Open Home tab on slow device','Skeleton loading animation visible briefly','','Not Run','','','No'),
    ('Progress','Step Indicator in Multi-Step Flows','Step progress shown in wizard flows','Medium','Resume creation wizard (if multi-step)','N/A','1. Enter resume creation\n2. Observe step indicator','Step 1 of N indicator visible','','Not Run','','','No'),
    ('Bottom Sheet','Draggable Sheet Has Drag Handle','Handle indicator at top of sheet','Medium','Any bottom sheet','N/A','1. Open backup sheet\n2. Check top','Pill/drag handle visible at top of sheet','','Not Run','','','No'),
    ('Snackbar','Snackbar Does Not Block FAB','Action snackbar does not hide FAB','Medium','Home tab; FAB visible','N/A','1. Trigger action that shows snackbar\n2. Check FAB','FAB remains accessible above snackbar','','Not Run','','','No'),
    ('Templates','Template Thumbnail Shows Correct Layout','Thumbnail matches actual template','High','Template picker','N/A','1. Open template picker\n2. Select template\n3. Compare thumb to render','Thumbnail visually matches rendered template','','Not Run','','','No'),
    ('Navigation','Active Tab Highlighted','Current tab in bottom nav visually active','High','All tabs','N/A','1. Switch between all tabs','Active tab icon/label highlighted','','Not Run','','','No'),
    ('Navigation','Back Arrow Visible on All Nested Screens','Back button always present on sub-screens','High','Navigate to nested screens','N/A','1. Open Editor, Settings, AI screens','Back arrow in AppBar on all sub-screens','','Not Run','','','No'),
    ('Error','Empty Image Placeholder','Broken image shows placeholder','Medium','Image set to invalid URL','N/A','1. Set profile photo to invalid reference','Placeholder shown; no broken image icon','','Not Run','','','No'),
    ('Animations','AI Response Text Streams Smoothly','AI text appears character by character or in chunks, not all at once','Low','AI feature; streaming enabled','N/A','1. Trigger AI response\n2. Observe text appearance','Text appears progressively; no jarring flash','','Not Run','','','No'),
    ('Density','Compact UI at Small Screen','Content accessible on 5-inch phone','High','5-inch phone','N/A','1. Open all key screens on 5-inch phone','All content accessible; no content hidden off-screen','','Not Run','','','No'),
    ('Visual','Plan Badges Visually Distinct','Free vs Premium badges clearly different','High','Subscription screen','N/A','1. Open upgrade screen','Premium badge clearly distinct from Free badge in colour and style','','Not Run','','','No'),
    ('Visual','Priority Label Colour-Coded in UI','Premium-locked items show visual indicator','High','AI tools list','N/A','1. Open AI tools with mixed lock states','Lock icon and "Premium" badge on locked items','','Not Run','','','No'),
    ('Layout','Landscape Layout on Large Phone','6.7-inch in landscape uses available width','Low','6.7-inch phone; landscape','N/A','1. Rotate large phone to landscape\n2. Open editor','Editor uses wider space; not tiny centered column','','Not Run','','','No'),
    ('Resume Card','Resume Card Shows Template Preview','Card shows thumbnail of current template','Medium','Home tab; resume with template','N/A','1. View resume cards\n2. Check template preview area','Template thumbnail or colour badge shown on card','','Not Run','','','No'),
    ('Settings','Settings Sections Visually Separated','APPEARANCE / AI / DATA / SUPPORT / ABOUT clearly divided','Medium','Settings screen','N/A','1. Open Settings\n2. Observe section headers','Section headers (APPEARANCE etc.) visually distinct with spacing','','Not Run','','','No'),
]

# ── EXTRA INTEGRATION ─────────────────────────────────────────────────────────
INTEGRATION_EXTRA = [
    ('ATS + Resume','ATS Analysis Reads Resume Model Correctly','ATS service receives correct resume data','High','Resume with experience and skills','N/A','1. Run ATS\n2. Verify analysis references correct resume fields','Analysis references actual resume content, not placeholder','','Not Run','','','Yes'),
    ('AI + Job Tracker','Job Tracker Status Feeds AI Context','Job status used as context in AI tools','Medium','Job in tracker with status; AI configured','N/A','1. Open job entry\n2. Generate AI tailored resume','AI prompt includes job status context','','Not Run','','','No'),
    ('Hive + Resume Quality','Resume Quality Score Uses Hive Data','Quality score reads from Hive, not in-memory only','High','Resume saved in Hive','N/A','1. Kill app\n2. Reopen\n3. Check quality score','Score persists from Hive data on reload','','Not Run','','','Yes'),
    ('Firebase + Crash Reporting','Firebase Crashlytics Records Crash','Crash report sent to Firebase on fatal error','Medium','Firebase configured; test crash triggered','N/A','1. Trigger test crash (debug mode)\n2. Check Firebase console','Crash report visible in Firebase Crashlytics','','Not Run','','','No'),
    ('Razorpay + Subscription','Razorpay Order ID Linked to Subscription','Order ID in Razorpay matches local subscription record','High','Razorpay payment complete','Payment: test order','1. Complete payment\n2. Check subscription detail (if visible)','Order ID linkage shown or verifiable in logs','','Not Run','','','No'),
    ('GoRouter + Auth Guard','Protected Routes Redirect to Login','Auth-required routes redirect unauthenticated users','Critical','No user logged in','N/A','1. Navigate to /resume-editor directly\n2. Observe','Redirected to login screen; not raw error page','','Not Run','','','Yes'),
    ('LibreTranslate + PDF','Translated Resume Exports Correctly','Translation service output used in PDF export','Medium','Translation service available; resume in English','Target: French','1. Translate to French\n2. Export PDF','French text in PDF; no garbled characters','','Not Run','','','No'),
    ('SkillSuggestions + Editor','Suggested Skills Inserted into Skills Section','Tapping suggested skill adds to resume skills','High','Skill suggestions visible','Skill: "Kotlin"','1. Open skill suggestions\n2. Tap "Add" on "Kotlin"','Kotlin added to resume skills section','','Not Run','','','Yes'),
    ('FreePlanService + UI','Premium Limit Countdown Shows Correct Remaining','AI use counter reflects actual usage','High','Free plan; 2 AI uses made','AI limit: 5','1. Check AI remaining count after 2 uses','Shows "3 free suggestions left"','','Not Run','','','Yes'),
    ('AppConfig + Feature Flags','Config-Driven Feature Enabled','Feature enabled via AppConfig key','High','SHOW_FEATURE_X=true in config','Feature X','1. Set config key\n2. Restart app\n3. Check feature visibility','Feature visible only when config key is true','','Not Run','','','Yes'),
    ('Resume Import + Storage','Imported Resume Persists to Hive','Imported resume saved to Hive correctly','High','File imported','resume.json','1. Import resume\n2. Kill app\n3. Reopen','Imported resume visible in resume list','','Not Run','','','Yes'),
    ('Version Service + UI','Version Info Shown in Settings from Service','Version service data used in About screen','High','App running','N/A','1. Open Settings → About','Version string matches AppVersionService output','','Not Run','','','Yes'),
    ('SyncStatusService + Backup UI','Sync Status Indicator Updated After Backup','Sync status in UI reflects last backup time','Medium','Backup completed','N/A','1. Backup\n2. Check sync status indicator','Last backup time updated in UI','','Not Run','','','Yes'),
    ('UserSession + AI Limits','AI Limits Tied to User Session','Switching accounts resets AI limit tracking','High','Two accounts with different usage','Account A: 3 uses; Account B: 0 uses','1. Login Account A (3 uses)\n2. Logout\n3. Login Account B\n4. Check AI limit','Account B shows 0/5 uses, not 3/5','','Not Run','','','Yes'),
    ('PDF + Image','Profile Photo Embedded Correctly in PDF','Profile photo shows in correct position in PDF','High','Profile photo set','JPG photo','1. Set profile photo\n2. Export PDF\n3. Open PDF','Photo in top-left or header of PDF at correct size','','Not Run','','','Yes'),
    ('Subscription + Feature Gate','Feature Gate Re-evaluates After Plan Change','Feature lock state re-checked after upgrade','Critical','Free plan; then upgrade','N/A','1. Check AI cover letter (locked)\n2. Upgrade\n3. Re-open cover letter','Cover letter accessible without app restart','','Not Run','','','Yes'),
    ('ResumeExport + Preview','Preview Uses Same Engine as PDF Export','Preview renders identically to final PDF','High','Resume with all sections','N/A','1. Preview resume\n2. Export PDF\n3. Compare','No visual discrepancies between preview and PDF','','Not Run','','','Yes'),
    ('DataDeletion + Auth','Data Deletion Does Not Log Out User','Delete data clears resume data but keeps auth','High','Logged in; resumes exist','N/A','1. Settings → Delete All Data\n2. Confirm','User still logged in; resume list empty','','Not Run','','','Yes'),
    ('AI + FreePlan','Free Plan AI Use Decremented Correctly','Each AI call decrements free limit counter','High','Free plan; 5 AI uses allowed','N/A','1. Make 3 AI calls\n2. Check counter','Counter shows 2 remaining','','Not Run','','','Yes'),
    ('JobTracker + CareerPath','Career Tools Tab State Preserved on Tab Switch','Job tracker scroll position preserved when switching tabs','Low','Job tracker with 10+ entries; scrolled down','N/A','1. Scroll job tracker to entry 8\n2. Switch to Home tab\n3. Return to Career Tools','Job tracker at same scroll position','','Not Run','','','No'),
]

# ── EXTRA SYSTEM ──────────────────────────────────────────────────────────────
SYSTEM_EXTRA = [
    ('Foreground Service','No Foreground Service in Background','App does not run foreground service when backgrounded','Medium','Battery settings; background app list','N/A','1. Background the app\n2. Check active services','No foreground service notification for the app','','Not Run','','','No'),
    ('File Storage','Scoped Storage Compliant (Android 10+)','App uses scoped storage APIs','High','Android 10+ device','N/A','1. Export PDF\n2. Check file location','PDF saved to MediaStore Downloads, not arbitrary path','','Not Run','','','No'),
    ('ANR Prevention','No ANR on Heavy Operations','No Application Not Responding on large data','High','Device with 50 resumes','N/A','1. Open Home with 50 resumes\n2. Observe','No ANR dialog; list loads within system limits','','Not Run','','','Yes'),
    ('Rotation','State Preserved on Orientation Change','Resume editor state maintained on rotation','High','Resume editor open','In-progress form data','1. Type in name field\n2. Rotate device\n3. Observe','Text retained; no form reset','','Not Run','','','Yes'),
    ('Multi-Window','App Works in Multi-Window Mode','Split-screen mode functional','Medium','Android 7+ device','N/A','1. Enable split screen\n2. Place app in one pane','App functional in split-screen pane','','Not Run','','','No'),
    ('Doze Mode','App Resumes After Doze','App functional after exiting Doze mode','Medium','Device left idle for Doze activation','N/A','1. Force Doze via adb\n2. Wake device\n3. Use app','App responsive; no stale state','','Not Run','','','No'),
    ('Background Kill','App Handles System Kill Gracefully','State restoration after system memory kill','High','Developer options: Don\'t keep activities','N/A','1. Open editor, partially filled\n2. Home → trigger memory pressure\n3. Return','App restores to editor; data intact from last save','','Not Run','','','Yes'),
    ('Locale Change','Locale Change at Runtime','Changing device language while app open','Medium','App running in English','Change to French','1. Change device language to French while app open','App handles locale change; no crash; restarts if needed','','Not Run','','','No'),
    ('Time Zone','Date Display Correct Across Time Zones','Dates shown in device local time','Medium','Device time zone changed','Time zone: JST UTC+9','1. Change time zone\n2. View job tracker dates','Dates displayed in new time zone','','Not Run','','','No'),
    ('Large Heap','App Does Not Request Large Heap','largeHeap not set in manifest (perf hygiene)','Low','Code review','AndroidManifest.xml','1. Check AndroidManifest.xml for largeHeap=true','largeHeap not set; app manages within standard heap','','Not Run','','','No'),
    ('File Descriptor Leak','No FD Leak on PDF Export','File descriptors not leaked on export','Medium','ADB: /proc/<pid>/fd count','N/A','1. Export PDF 5 times\n2. Check FD count','FD count does not grow unboundedly','','Not Run','','','No'),
    ('Accessibility Service','App Works with Accessibility Service Active','No crash with accessibility services running','Medium','TalkBack or Switch Access active','N/A','1. Enable accessibility service\n2. Run core flows','App functional with accessibility service active','','Not Run','','','No'),
    ('Dark Theme System','System Dark Mode Auto-Applies When Theme = System','App dark when system goes dark in System mode','High','App theme set to System Default','Device dark mode','1. Set app theme: System\n2. Toggle device dark mode','App switches theme to match system','','Not Run','','','Yes'),
    ('Notification Channel','Notification Channel Created Correctly','Android 8+ notification channel set up','High','Android 8+ device','N/A','1. Open device notification settings for app','Channel "Resume Builder" visible with correct settings','','Not Run','','','No'),
    ('Instant App','Cold Start on First Install','App starts clean on first-ever install','High','Newly installed device','N/A','1. Install fresh APK\n2. Open','Splash → onboarding; no crash; no stale data artifacts','','Not Run','','','Yes'),
]

# ── EXTRA SECURITY ────────────────────────────────────────────────────────────
SECURITY_EXTRA = [
    ('Input','HTML Injection in PDF via Job Tracker Notes','HTML tags in notes do not execute','High','Job tracker note field','Note: <b>bold</b><script>…</script>','1. Add HTML in job notes\n2. View note','Note shown as raw text; HTML not rendered or executed','','Not Run','','','Yes'),
    ('Auth','Concurrent Session Management','Two devices using same account behave predictably','Medium','Same account on two devices','N/A','1. Login on Device A\n2. Login on Device B\n3. Use both','Both sessions functional; no data conflict or silent logout','','Not Run','','','No'),
    ('Auth','OTP Not Reusable','Used OTP cannot be replayed','Critical','OTP used once','Same OTP','1. Use OTP to login\n2. Attempt login again with same OTP','OTP rejected on second use','','Not Run','','','Yes'),
    ('Dependency','No Known Vulnerable Dependencies','All pub dependencies free of critical CVEs','High','pub outdated; security advisory','pubspec.lock','1. Run dart pub outdated\n2. Check for CVE advisories','No dependencies with known critical CVEs in use','','Not Run','','','No'),
    ('Certificate','SSL/TLS Version','TLS 1.2+ required; TLS 1.0/1.1 not accepted','High','Network inspector','N/A','1. Attempt connection with TLS 1.0 via proxy','Connection rejected; TLS 1.2 minimum enforced','','Not Run','','','No'),
    ('Privacy','GDPR Data Deletion','User can request full data deletion','High','Settings; Supabase data','N/A','1. Settings → Delete All Data\n2. Confirm','All local data deleted; cloud data deletion triggered','','Not Run','','','No'),
    ('Privacy','Analytics Opt-Out Respected','No analytics sent when user opts out','Medium','Analytics opt-out setting (if present)','N/A','1. Opt out of analytics\n2. Use app\n3. Check network traffic','No analytics events sent after opt-out','','Not Run','','','No'),
    ('Keystore','Signing Key Not in Source Code','Upload keystore and credentials not committed','Critical','Git repo check','N/A','1. Check .gitignore for *.jks\n2. Search git history','No .jks files or key passwords in git history','','Not Run','','','No'),
    ('Proguard','ProGuard/R8 Obfuscation Verified','Release APK class names obfuscated','High','Decompile release APK with jadx','N/A','1. Decompile\n2. Inspect class structure','Meaningful class/method names absent; only a, b, c…','','Not Run','','','No'),
    ('Deep Link','Deep Link Cannot Open Sensitive Routes Without Auth','Unauthenticated deep links to protected screens blocked','High','Not logged in','Deep link: /backup-sync','1. Trigger /backup-sync deep link when not logged in','Redirected to login; backup screen not shown','','Not Run','','','Yes'),
    ('Clipboard','Sensitive Data Not Auto-Copied to Clipboard','OTP or token not auto-filled to clipboard','Medium','OTP screen','N/A','1. Receive OTP\n2. Check clipboard content','OTP not silently copied to clipboard by app','','Not Run','','','No'),
    ('Screenshot','Secure Fields Prevent Screenshot on Android','Flutter secure storage fields not in screenshots (if enabled)','Low','Android Recent Apps','N/A','1. Open page with API key setting\n2. Take screenshot via Recent Apps','Blank screenshot if FLAG_SECURE is set for sensitive screens','','Not Run','','','No'),
    ('Network','No Cleartext Traffic','Manifest cleartext traffic blocked','Critical','AndroidManifest review','N/A','1. Check android:usesCleartextTraffic in manifest','usesCleartextTraffic=false in manifest or network_security_config','','Not Run','','','No'),
    ('Data Backup','Android Backup Disabled for Sensitive Data','android:allowBackup=false or restricted','High','AndroidManifest review','N/A','1. Check allowBackup setting','Sensitive Hive and secure storage excluded from Android backup','','Not Run','','','No'),
    ('Injection','OS Command Injection via File Name','Custom PDF filename cannot execute OS commands','High','Export PDF with custom name','Filename: "$(reboot).pdf"','1. Set PDF filename to injection string\n2. Export','File created with literal filename; no OS command executed','','Not Run','','','Yes'),
    ('Obfuscation','Debug Logs Stripped in Release Build','No debug print statements in release APK','High','Decompile or logcat in release','N/A','1. Install release APK\n2. Check Logcat','No debug messages (tag "flutter") in release log','','Not Run','','','Yes'),
]

# ── EXTRA BOUNDARY ────────────────────────────────────────────────────────────
BOUNDARY_EXTRA = [
    ('Resume Editor','Date Year – 1900','Very old graduation year accepted','Low','Education section','Year: 1900','1. Enter graduation year 1900\n2. Save','Year saved without error','','Not Run','','','Yes'),
    ('Resume Editor','Date Year – 2099','Future date accepted for projected end','Low','Experience section','End year: 2099','1. Enter end year 2099\n2. Save','Year saved; shown in preview','','Not Run','','','Yes'),
    ('AI','AI Prompt – Exactly 5 Characters','Very short prompt processed','Medium','AI enhancer','Prompt: "React"','1. Enter 5-char prompt\n2. Enhance','Response or helpful error; no crash','','Not Run','','','Yes'),
    ('Sync','Sync Code – 50 Characters (Mid Boundary)','50-char sync code accepted','Medium','Backup sheet','Code: 50-char alphanumeric','1. Set 50-char code\n2. Backup','Backup succeeds','','Not Run','','','Yes'),
    ('Sync','Sync Code – Special Characters Rejected','Sync code with spaces/symbols shows error','Medium','Backup sheet','Code: "bad code!!"','1. Enter code with spaces and "!!"\n2. Save','Validation error: allowed characters only','','Not Run','','','Yes'),
    ('Resume','Resume Name – All Spaces Rejected','Name with only spaces not accepted','Medium','Resume creation','Name: "     "','1. Enter all-space name\n2. Tap Create','Validation error: "Name cannot be blank"','','Not Run','','','Yes'),
    ('Job Tracker','Job Notes – 500 Characters','Long job note accepted','Low','Job tracker note field','500-char note','1. Enter 500-char note\n2. Save','Note saved; fully visible on scroll','','Not Run','','','Yes'),
    ('Subscription','Free Trial – 0 Days Remaining','Trial expired message at zero days','High','Trial with 0 days left','N/A','1. Open app on last day of trial (expired)','Plan shows as expired; premium features locked','','Not Run','','','Yes'),
]

BOUNDARY.extend(BOUNDARY_EXTRA)

# Extend base lists with extras
FUNCTIONAL.extend(FUNCTIONAL_EXTRA)
REGRESSION.extend(REGRESSION_EXTRA)
UI_UX.extend(UI_UX_EXTRA)
INTEGRATION.extend(INTEGRATION_EXTRA)
SYSTEM.extend(SYSTEM_EXTRA)
SECURITY.extend(SECURITY_EXTRA)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET DATA MAPPING
# ══════════════════════════════════════════════════════════════════════════════
SHEET_DATA = {
    'Functional':           FUNCTIONAL,
    'End-to-End E2E':       E2E,
    'Smoke':                SMOKE,
    'Sanity':               SANITY,
    'Regression':           REGRESSION,
    'UI-UX':                UI_UX,
    'API':                  API,
    'Integration':          INTEGRATION,
    'System':               SYSTEM,
    'Security':             SECURITY,
    'Performance':          PERFORMANCE,
    'Compatibility':        COMPATIBILITY,
    'Accessibility':        ACCESSIBILITY,
    'Negative':             NEGATIVE,
    'Boundary':             BOUNDARY,
    'Subscription-Billing': SUBSCRIPTION,
    'Backup-Sync':          BACKUP_SYNC,
    'Cross-Platform':       CROSS_PLATFORM,
}


# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Cover / Index sheet ───────────────────────────────────────────────────
    ws_cover = wb.create_sheet('INDEX')
    ws_cover.sheet_properties.tabColor = 'FF2F5496'
    ws_cover.column_dimensions['A'].width = 30
    ws_cover.column_dimensions['B'].width = 20
    ws_cover.column_dimensions['C'].width = 15
    ws_cover.column_dimensions['D'].width = 40

    cover_hdr_fill = PatternFill('solid', fgColor='FF2F5496')
    cover_hdr_font = Font(bold=True, color='FFFFFFFF', size=11)

    ws_cover['A1'] = 'Resume Builder App – Test Case Workbook'
    ws_cover['A1'].font = Font(bold=True, size=16, color='FF2F5496')
    ws_cover.merge_cells('A1:D1')
    ws_cover['A1'].alignment = Alignment(horizontal='center')

    ws_cover['A2'] = f'Generated on: 2026-06-07'
    ws_cover.merge_cells('A2:D2')
    ws_cover['A2'].font = Font(italic=True, size=10, color='FF595959')
    ws_cover['A2'].alignment = Alignment(horizontal='center')

    ws_cover.append([])  # blank row

    headers = ['Testing Type', 'Sheet Name', 'Test Count', 'Colour Code']
    ws_cover.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws_cover.cell(row=4, column=col_idx, value=h)
        cell.fill = cover_hdr_fill
        cell.font = cover_hdr_font
        cell.alignment = Alignment(horizontal='center')

    total = 0
    for i, (sheet_name, tab_color, prefix) in enumerate(SHEETS):
        data = SHEET_DATA.get(sheet_name, [])
        count = len(data)
        total += count
        row = [sheet_name, sheet_name, count, f'#{tab_color[2:]}']
        ws_cover.append(row)
        r = ws_cover.max_row
        # colour swatch
        ws_cover.cell(r, 1).fill = PatternFill('solid', fgColor=tab_color)
        ws_cover.cell(r, 1).font = Font(bold=True,
                                         color='FFFFFFFF' if _luminance(tab_color) < 0.6 else 'FF1F1F1F')

    ws_cover.append([])
    ws_cover.append(['TOTAL TEST CASES', '', total, ''])
    total_row = ws_cover.max_row
    for col in range(1, 5):
        ws_cover.cell(total_row, col).font = Font(bold=True, size=11)
    ws_cover.cell(total_row, 3).fill = PatternFill('solid', fgColor='FF70AD47')
    ws_cover.cell(total_row, 3).font = Font(bold=True, color='FFFFFFFF', size=11)

    # ── Data sheets ───────────────────────────────────────────────────────────
    for sheet_name, tab_color, prefix in SHEETS:
        data = SHEET_DATA.get(sheet_name, [])
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab_color

        # Column widths
        for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # Freeze panes
        ws.freeze_panes = 'A2'

        # Header row
        hdr_fill, hdr_font = make_header_style(tab_color)
        for col_idx, (col_name, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border()

        ws.row_dimensions[1].height = 30

        # Data rows
        for row_idx, record in enumerate(data, 2):
            tc_id = f'{prefix}-{str(row_idx - 1).zfill(3)}'
            # record: (Module, Scenario, Description, Priority, Preconditions,
            #           Test Data, Test Steps, Expected Result, Automation)
            row_fill = ALT_ROW_FILL if row_idx % 2 == 0 else WHITE_FILL
            row_data = [tc_id] + list(record)

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(
                    vertical='top', wrap_text=True)
                cell.border = thin_border()

                # Priority colour (col 5)
                if col_idx == 5:
                    p = str(value)
                    cell.fill = priority_fill(p)
                    cell.font = Font(bold=True, size=9,
                                     color='FF1F1F1F' if _luminance(PRIORITY_COLOURS.get(p, 'FFFFFFFF')) > 0.5 else 'FFFFFFFF')
                # Status colour (col 11)
                elif col_idx == 11:
                    cell.fill = PatternFill('solid',
                                            fgColor=STATUS_COLOURS.get(str(value), 'FFD9D9D9'))
                    cell.font = Font(bold=False, size=9)
                # Automation (col 14)
                elif col_idx == 14:
                    cell.fill = PatternFill('solid',
                                            fgColor='FFD9EAD3' if value == 'Yes' else 'FFFFF2CC')
                    cell.font = Font(bold=True, size=9)
                else:
                    cell.fill = row_fill

            ws.row_dimensions[row_idx].height = 60

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

    return wb


def _luminance(hex_color: str) -> float:
    hex_color = hex_color.lstrip('#')
    if len(hex_color) == 8:
        hex_color = hex_color[2:]
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    return (0.299 * r + 0.587 * g + 0.114 * b) / 255


if __name__ == '__main__':
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    print('Building workbook …')
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    total = sum(len(v) for v in SHEET_DATA.values())
    print(f'Done. {total} test cases across {len(SHEETS)} sheets.')
    print(f'Saved → {os.path.abspath(OUTPUT_PATH)}')
